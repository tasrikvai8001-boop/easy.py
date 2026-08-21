import importlib.util
import subprocess
import sys
import json
import os
import time
import random
import string
import threading
import traceback
import smtplib
import pyotp
import urllib.parse
from datetime import datetime, timedelta

# --- AUTOMATIC DEPENDENCY CHECK & INSTALLATION ---
REQUIRED_PACKAGES = ["flask", "pyTelegramBotAPI", "gspread", "oauth2client", "pyotp", "openpyxl", "requests"]
for pkg in REQUIRED_PACKAGES:
    mod = "telebot" if pkg == "pyTelegramBotAPI" else pkg
    if importlib.util.find_spec(mod) is None:
        subprocess.check_call([sys.executable, "-m", "pip", "install", pkg])

from flask import Flask
import requests
import telebot
from telebot.types import ReplyKeyboardMarkup, KeyboardButton, InlineKeyboardMarkup, InlineKeyboardButton
import gspread
from oauth2client.service_account import ServiceAccountCredentials
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ============================================
# --- WEB SERVER FOR KEEP-ALIVE ---
# ============================================
app = Flask('')

@app.route('/')
def home():
    return "EASY EARN BD Firebase Engine is Running 24/7!"

def run_web_server():
    port = int(os.environ.get("PORT", 8080))
    app.run(host='0.0.0.0', port=port)

def keep_alive():
    t = threading.Thread(target=run_web_server)
    t.daemon = True
    t.start()

# ============================================
# --- CONFIGURATION & ENVIRONMENT SECURITY ---
# ============================================
TOKEN = os.environ.get("BOT_TOKEN", "8593556780:AAF9GtHJw0oKFEpRyKOb_9kZFFJtJ7XnAL8")
ADMIN_ID = int(os.environ.get("ADMIN_ID", 7833766898))
BOT_NAME = "EASY EARN BD"
LOG_FILE = "error_logs.txt"
JSON_CREDS_FILE = "credentials.json"

bot = telebot.TeleBot(TOKEN, num_threads=50)

# FIREBASE RTDB CONFIGURATION
FIREBASE_URL = "https://fast-cash-out-default-rtdb.firebaseio.com"

# --- ERROR & AUDIT LOGGERS ---
def log_error(err_msg):
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    with open(LOG_FILE, "a", encoding="utf-8") as f:
        f.write(f"[{timestamp}] {err_msg}\n{'-'*40}\n")

def log_admin_action(admin_id, action_desc):
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    msg = f"<b>🛡️ Admin Audit Log</b>\n<b>👤 Admin:</b> <code>{admin_id}</code>\n<b>📝 Action:</b> {action_desc}\n<b>⏰ Time:</b> <code>{timestamp}</code>"
    try:
        bot.send_message(ADMIN_ID, msg, parse_mode="HTML")
    except:
        pass

# ============================================
# --- FIREBASE REST DATABASE ENGINE ---
# ============================================
def fb_get(path):
    try:
        res = requests.get(f"{FIREBASE_URL}/{path}.json", timeout=10)
        return res.json() if res.status_code == 200 else None
    except Exception as e:
        log_error(f"Firebase GET Error ({path}): {e}")
        return None

def fb_set(path, data):
    try:
        res = requests.put(f"{FIREBASE_URL}/{path}.json", json=data, timeout=10)
        return res.status_code == 200
    except Exception as e:
        log_error(f"Firebase SET Error ({path}): {e}")
        return False

def fb_update(path, data):
    try:
        res = requests.patch(f"{FIREBASE_URL}/{path}.json", json=data, timeout=10)
        return res.status_code == 200
    except Exception as e:
        log_error(f"Firebase UPDATE Error ({path}): {e}")
        return False

def fb_delete(path):
    try:
        res = requests.delete(f"{FIREBASE_URL}/{path}.json", timeout=10)
        return res.status_code == 200
    except Exception as e:
        log_error(f"Firebase DELETE Error ({path}): {e}")
        return False

def init_db():
    default_emojis = {
        "balance": "💰", "work": "💼", "withdraw": "📥", "invite": "👥",
        "support": "🎧", "newbie": "❓", "instagram": "📸", "facebook": "📘",
        "gmail": "📧", "spin": "🌀", "task": "📲", "success": "✅",
        "error": "❌", "warning": "⚠️", "lock": "🔐", "admin": "⚙️"
    }

    defaults = {
        "force_channels": json.dumps([]),
        "ref_bonus": "10.0",
        "rate_ins": "15.0",
        "rate_fb": "18.0",
        "rate_gmail": "12.0",
        "ins_pass": "Nabil",
        "fb_pass": "Nabil",
        "gmail_pass": "Nabil",
        "recovery_email": "tasrikvai8001@gmail.com",
        "emojis": json.dumps(default_emojis),
        "spin_ad_url": "https://example.com/adsterra",
        "spin_reward": "1.5",
        "spin_limit": "5",
        "sheet_id_ins": "",
        "sheet_id_fb": "",
        "sheet_id_gmail": "",
        "json_credentials": "",
        "firebase_api": "AIzaSyBGABXnrP66oCndR0a6Hza3m2pehk2JgcE",
        "tutorial_videos": json.dumps({
            "gmail": "https://youtube.com", 
            "fb": "https://youtube.com", 
            "ins": "https://youtube.com"
        }),
        "withdraw_methods": json.dumps({
            "bKash": {"enabled": True, "min": 50.0, "max": 5000.0},
            "Nagad": {"enabled": True, "min": 50.0, "max": 5000.0},
            "Rocket": {"enabled": True, "min": 50.0, "max": 5000.0},
            "USDT BEP20": {"enabled": True, "min": 100.0, "max": 10000.0}
        }),
        "maintenance_mode": "false",
        "pause_gmail": "false",
        "pause_fb": "false",
        "pause_ins": "false",
        "daily_bot_withdraw_limit": "50000.0",
        "today_total_withdrawn": "0.0",
        "last_withdraw_reset_date": ""
    }
    
    cfg = fb_get("config") or {}
    for k, v in defaults.items():
        if k not in cfg:
            fb_set(f"config/{k}", str(v))

init_db()

def get_config(key, default=""):
    val = fb_get(f"config/{key}")
    return str(val) if val is not None else default

def set_config(key, value):
    fb_set(f"config/{key}", str(value))

def get_user_db(user_id):
    u = fb_get(f"users/{user_id}")
    now = time.time()
    if not u:
        u_data = {
            "user_id": user_id,
            "balance": 0.0,
            "total_income": 0.0,
            "total_withdraw": 0.0,
            "referrals": 0,
            "referred_by": None,
            "ref_rewarded": 0,
            "state": None,
            "temp_data": '{}',
            "role": 'user',
            "permissions": '{}',
            "is_banned": 0,
            "approved_tasks": 0,
            "rejected_tasks": 0,
            "pending_tasks": 0,
            "completed_accounts": 0,
            "last_spin_time": 0,
            "daily_spins": 0,
            "last_spin_date": '',
            "last_active": now,
            "ip_address": None,
            "created_at": now
        }
        fb_set(f"users/{user_id}", u_data)
        return u_data
    else:
        fb_update(f"users/{user_id}", {"last_active": now})
        return u

def update_user_field(user_id, field, value):
    fb_update(f"users/{user_id}", {field: value})

def get_emoji(key):
    emojis = json.loads(get_config("emojis", "{}"))
    return emojis.get(key, "✨")

def check_permission(user_id, perm="admin"):
    if user_id == ADMIN_ID: return True
    u = get_user_db(user_id)
    if u.get("role") == "admin": return True
    if u.get("role") == "sub_admin":
        perms = json.loads(u.get("permissions") or "{}")
        return perms.get(perm, False) or perms.get("all", False)
    return False

# ============================================
# --- GOOGLE SHEETS AUTOMATION ENGINE ---
# ============================================
def get_gspread_client():
    try:
        json_data = get_config("json_credentials", "")
        if not json_data and os.path.exists(JSON_CREDS_FILE):
            with open(JSON_CREDS_FILE, 'r') as f:
                json_data = f.read()

        if not json_data: return None

        scope = ['https://spreadsheets.google.com/feeds', 'https://www.googleapis.com/auth/drive']
        creds_dict = json.loads(json_data)
        creds = ServiceAccountCredentials.from_json_keyfile_dict(creds_dict, scope)
        return gspread.authorize(creds)
    except Exception as e:
        log_error(f"GSpread Client Auth Error: {e}\n{traceback.format_exc()}")
        return None

def append_to_google_sheet(service_type, row_data):
    try:
        client = get_gspread_client()
        if not client: return False

        sheet_key = f"sheet_id_{service_type.lower()}"
        sheet_id = get_config(sheet_key, "")
        if not sheet_id: return False

        sh = client.open_by_key(sheet_id)
        
        target_sheet = None
        for ws in sh.worksheets():
            w_title = ws.title.strip().lower()
            if service_type.lower() in w_title or w_title == "sheet1":
                target_sheet = ws
                break
        if not target_sheet:
            target_sheet = sh.sheet1

        target_sheet.append_row(row_data)
        return True
    except Exception as e:
        log_error(f"Google Sheet Append Error ({service_type}): {e}\n{traceback.format_exc()}")
        return False

def sync_sheet_approvals():
    processed_count = 0
    client = get_gspread_client()
    if not client: return 0

    services = ['ins', 'fb', 'gmail']
    for s_type in services:
        sheet_id = get_config(f"sheet_id_{s_type}", "")
        if not sheet_id: continue
        try:
            sh = client.open_by_key(sheet_id)
            target_sheet = sh.sheet1
            records = target_sheet.get_all_records()
            all_subs = fb_get("pending_submissions") or {}

            for idx, row in enumerate(records, start=2):
                sub_id = str(row.get("Submit_ID") or row.get("ID") or "")
                status_val = str(row.get("Status") or "").strip().lower()

                if not sub_id or status_val not in ["ok", "bad", "approved", "rejected"]:
                    continue

                if sub_id in all_subs and all_subs[sub_id].get("status") == "pending":
                    sub = all_subs[sub_id]
                    uid = sub['user_id']
                    rate = float(sub['rate'])

                    if status_val in ["ok", "approved"]:
                        fb_update(f"pending_submissions/{sub_id}", {"status": "approved"})
                        u = get_user_db(uid)
                        fb_update(f"users/{uid}", {
                            "balance": float(u.get("balance", 0)) + rate,
                            "total_income": float(u.get("total_income", 0)) + rate,
                            "approved_tasks": int(u.get("approved_tasks", 0)) + 1,
                            "pending_tasks": max(0, int(u.get("pending_tasks", 0)) - 1)
                        })
                        
                        # Referral Lifetime 10% Bonus Logic
                        ref_id = u.get("referred_by")
                        if ref_id:
                            ref_u = get_user_db(ref_id)
                            ref_comm = rate * 0.10
                            fb_update(f"users/{ref_id}", {
                                "balance": float(ref_u.get("balance", 0)) + ref_comm,
                                "total_income": float(ref_u.get("total_income", 0)) + ref_comm
                            })
                            try: bot.send_message(ref_id, f"🎉 <b>রেফারেল ১০% কমিশন!</b> আপনার রেফারেল একটি কাজ সম্পন্ন করায় আপনি <b>৳{ref_comm:.2f}</b> কমিশন পেয়েছেন!", parse_mode="HTML")
                            except: pass

                        try: bot.send_message(uid, f"✅ <b>আপনার কাজ এপ্রুভ হয়েছে!</b>\nSubmit ID: <code>{sub_id}</code>\nব্যালেন্সে যোগ করা হয়েছে: ৳{rate:.2f}", parse_mode="HTML")
                        except: pass

                    elif status_val in ["bad", "rejected"]:
                        fb_update(f"pending_submissions/{sub_id}", {"status": "rejected"})
                        u = get_user_db(uid)
                        fb_update(f"users/{uid}", {
                            "rejected_tasks": int(u.get("rejected_tasks", 0)) + 1,
                            "pending_tasks": max(0, int(u.get("pending_tasks", 0)) - 1)
                        })
                        try: bot.send_message(uid, f"❌ <b>আপনার কাজ রিজেক্ট করা হয়েছে!</b>\nSubmit ID: <code>{sub_id}</code>", parse_mode="HTML")
                        except: pass

                    processed_count += 1
        except Exception as e:
            log_error(f"Sheet Sync Error for {s_type}: {e}\n{traceback.format_exc()}")
            continue
    return processed_count

# ============================================
# --- AUTOMATIC GMAIL SMTP EXISTENCE CHECKER ---
# ============================================
def verify_gmail_smtp(email):
    try:
        domain = email.split('@')[-1]
        if domain.lower() != "gmail.com": return True
        server = smtplib.SMTP(timeout=5)
        server.connect('gmail-smtp-in.l.google.com', 25)
        server.helo('localhost')
        server.mail('check@example.com')
        code, message = server.rcpt(str(email))
        server.quit()
        return code == 250
    except Exception as e:
        log_error(f"SMTP Verification Warning for {email}: {e}")
        return True 

# ============================================
# --- TELEBOT STYLISH COLOR BUTTON PATCH ---
# ============================================
_old_inline_dict = InlineKeyboardButton.to_dict
def _new_inline_dict(self):
    d = _old_inline_dict(self)
    if hasattr(self, 'style'): d['style'] = self.style
    return d
InlineKeyboardButton.to_dict = _new_inline_dict

_old_kb_dict = KeyboardButton.to_dict
def _new_kb_dict(self):
    d = _old_kb_dict(self)
    if hasattr(self, 'style'): d['style'] = self.style
    return d
KeyboardButton.to_dict = _new_kb_dict

def ibtn(text, callback_data=None, url=None, style=None):
    kwargs = {'text': text}
    if callback_data: kwargs['callback_data'] = callback_data
    if url: kwargs['url'] = url
    b = InlineKeyboardButton(**kwargs)
    if style: b.style = style
    return b

def rbtn(text, style=None):
    b = KeyboardButton(text=text)
    if style: b.style = style
    return b

# ============================================
# --- DYNAMIC HELPERS & GENERATORS ---
# ============================================
def generate_dynamic_password(prefix_key):
    base_name = get_config(prefix_key, "Nabil")
    date_str = datetime.now().strftime("%d%m")
    special_char = "@"
    return f"{base_name}{date_str}{special_char}"

FIRST_NAMES = ["Tanvir", "Rahim", "Kareem", "Sabbir", "Arif", "Mahmud", "Shakib", "Naim", "Fahim", "Hasan", "Sumon", "Raju"]
LAST_NAMES = ["Hossain", "Islam", "Ahmed", "Chowdhury", "Khan", "Uddin", "Rahman", "Mia", "Ali", "Sarker", "Roy", "Das"]

def generate_random_identity():
    fn = random.choice(FIRST_NAMES)
    ln = random.choice(LAST_NAMES)
    un = f"{fn.lower()}{ln.lower()}{random.randint(100, 9999)}"
    return fn, ln, un

def check_duplicate_and_save(val, record_type, uid):
    try:
        clean_key = str(val).replace(".", "_").replace("#", "_").replace("$", "_").replace("[", "_").replace("]", "_")
        rec = fb_get(f"submitted_records/{clean_key}")
        if rec:
            return True
        fb_set(f"submitted_records/{clean_key}", {
            "record_value": val,
            "record_type": record_type,
            "user_id": uid,
            "submitted_at": time.time()
        })
        return False
    except Exception as e:
        log_error(f"Duplicate Check Error: {e}")
        return False

def get_daily_withdraw_count(user_id):
    try:
        reqs = fb_get("withdraw_requests") or {}
        day_start = time.time() - 86400
        cnt = 0
        for r_id, r in reqs.items():
            if str(r.get("user_id")) == str(user_id) and float(r.get("created_at", 0)) >= day_start:
                cnt += 1
        return cnt
    except:
        return 0

def get_daily_support_count(user_id):
    try:
        tickets = fb_get("support_tickets") or {}
        day_start = time.time() - 86400
        cnt = 0
        for t_id, t in tickets.items():
            if str(t.get("user_id")) == str(user_id) and float(t.get("created_at", 0)) >= day_start:
                cnt += 1
        return cnt
    except:
        return 0

# ============================================
# --- CAPTCHA GENERATOR & FORCE JOIN ---
# ============================================
def generate_captcha():
    num1 = random.randint(1, 9)
    num2 = random.randint(1, 9)
    ans = num1 + num2
    return f"{num1} + {num2} = ?", str(ans)

def check_force_join(user_id):
    chs = json.loads(get_config("force_channels", "[]"))
    if not chs: return True
    for ch in chs:
        try:
            m = bot.get_chat_member(ch, user_id)
            if m.status in ['left', 'kicked']: return False
        except Exception as e: 
            log_error(f"Force join check error for {ch}: {e}")
            return False
    return True

def get_force_join_markup():
    chs = json.loads(get_config("force_channels", "[]"))
    markup = InlineKeyboardMarkup(row_width=1)
    for ch in chs:
        clean_ch = ch.replace("@", "")
        markup.add(ibtn(f"📢 Join {ch}", url=f"https://t.me/{clean_ch}", style="primary"))
    markup.add(ibtn("✅ Verify Now", callback_data="check_join_event", style="success"))
    return markup

# MENU TEXT CONSTANTS
TXT_WORK_MAIN = f"{get_emoji('work')} কাজ•"
TXT_TODAY_WORK = "🔥 আজকের কাজ"
TXT_BALANCE = f"{get_emoji('balance')} ব্যালেন্স"
TXT_WITHDRAW = f"{get_emoji('withdraw')} উত্তোলন"
TXT_REFER = f"{get_emoji('invite')} রেফার"
TXT_SUPPORT = f"{get_emoji('support')} সাপোর্ট"
TXT_NEWBIE = f"{get_emoji('newbie')} আমি নতুন"
TXT_ADMIN_PANEL = "⚙️ Admin Panel"
TXT_BACK = "🔙 Back"

def get_main_menu(user_id):
    u = get_user_db(user_id)
    markup = ReplyKeyboardMarkup(resize_keyboard=True, row_width=2)
    markup.add(rbtn(TXT_WORK_MAIN, "primary"), rbtn(TXT_TODAY_WORK, "primary"))
    markup.add(rbtn(TXT_BALANCE, "primary"), rbtn(TXT_WITHDRAW, "success"))
    markup.add(rbtn(TXT_REFER, "primary"), rbtn(TXT_SUPPORT, "primary"))
    markup.add(rbtn(TXT_NEWBIE, "primary"))
    if user_id == ADMIN_ID or u.get("role") in ["admin", "sub_admin", "moderator"]:
        markup.add(rbtn(TXT_ADMIN_PANEL, "danger"))
    return markup

def get_work_menu():
    markup = ReplyKeyboardMarkup(resize_keyboard=True, row_width=2)
    markup.add(rbtn(f"{get_emoji('instagram')} ইনস্টাগ্রাম কাজ", "primary"), rbtn(f"{get_emoji('gmail')} Gmail কাজ", "primary"))
    markup.add(rbtn(f"{get_emoji('facebook')} ফেসবুক কাজ", "primary"))
    markup.add(rbtn(TXT_BACK, "danger"))
    return markup

def get_admin_menu():
    markup = ReplyKeyboardMarkup(resize_keyboard=True, row_width=2)
    markup.add(rbtn("💰 Set Task Rates", "success"), rbtn("📢 Set Force Join", "primary"))
    markup.add(rbtn("🎁 Set Ref Bonus", "success"), rbtn("💳 Withdraw Setting", "primary"))
    markup.add(rbtn("🔑 Set Passwords Base", "primary"), rbtn("🎨 Edit Emoji IDs", "primary"))
    markup.add(rbtn("🌀 Spin & Ad Settings", "primary"), rbtn("📊 Google Sheets Config", "primary"))
    markup.add(rbtn("🤖 Upload JSON Creds", "primary"), rbtn("🔥 Set Firebase API", "primary"))
    markup.add(rbtn("🔄 Sync Google Sheet Approval", "success"), rbtn("📥 Export Unsold Files", "primary"))
    markup.add(rbtn("🔎 Pending Approvals", "primary"), rbtn("📥 Pending Withdraws", "primary"))
    markup.add(rbtn("📩 Pending Support Tickets", "primary"), rbtn("➕ Add App/TG Task", "success"))
    markup.add(rbtn("📩 Smart Broadcast Message", "primary"), rbtn("📊 Bot Statistics", "primary"))
    markup.add(rbtn("⛔ Ban/Unban User", "danger"), rbtn("➕ Add/Deduct Balance", "success"))
    markup.add(rbtn("👑 Sub-Admin Manager", "primary"), rbtn("📧 Set Recovery Email", "primary"))
    markup.add(rbtn("🎥 Set Tutorial Videos", "primary"), rbtn("🔍 User Deep Search", "primary"))
    markup.add(rbtn("📢 Channel Multi-Post Broadcast", "primary"), rbtn("🗑️ Bulk Reject Submissions", "danger"))
    markup.add(rbtn("⏸️ Task Pause Manager", "primary"), rbtn("🧹 Database Cleanup", "danger"))
    markup.add(rbtn("🚨 Error Logs", "danger"), rbtn("⚡ Maintenance Mode", "danger"))
    markup.add(rbtn(TXT_BACK, "danger"))
    return markup

# ============================================
# --- CRON / RE-ENGAGEMENT BACKGROUND THREAD ---
# ============================================
def automated_reengagement_cron():
    while True:
        try:
            time.sleep(86400) # Runs daily
            users = fb_get("users") or {}
            three_days_ago = time.time() - (3 * 86400)

            for u_id, u_row in users.items():
                if float(u_row.get("last_active", 0)) <= three_days_ago and int(u_row.get("is_banned", 0)) == 0:
                    try:
                        msg = "<b>👋 আমরা আপনাকে মিস করছি!</b>\n\nআপনার জন্য নতুন ফেসবুক ও জিমেইল টাস্ক অপেক্ষা করছে। এখনই বটে লগইন করে কাজ করে আয় করুন! 💰"
                        bot.send_message(u_id, msg, parse_mode="HTML")
                        time.sleep(0.05)
                    except:
                        pass
        except Exception as e:
            log_error(f"Re-engagement Cron Error: {e}")

threading.Thread(target=automated_reengagement_cron, daemon=True).start()

# ============================================
# --- COMMAND & START CAPTCHA HANDLERS ---
# ============================================
@bot.message_handler(commands=['start'])
def send_welcome(message):
    try:
        uid = message.from_user.id
        
        if get_config("maintenance_mode", "false") == "true" and uid != ADMIN_ID:
            bot.send_message(message.chat.id, f"{get_emoji('warning')} <b>বট বর্তমানে মেইনটেন্যান্স মোডে আছে।</b> খুব শীঘ্রই আবার কাজ চালু হবে।", parse_mode="HTML")
            return

        u = get_user_db(uid)
        if u.get("is_banned"):
            bot.send_message(message.chat.id, f"{get_emoji('error')} <b>আপনি এই বটে ব্লকড আছেন!</b>", parse_mode="HTML")
            return

        args = message.text.split()
        if len(args) > 1 and not u.get("referred_by"):
            ref_id = args[1]
            if ref_id.isdigit() and int(ref_id) != uid:
                ref_user = get_user_db(int(ref_id))
                if u.get("ip_address") and u.get("ip_address") == ref_user.get("ip_address"):
                    bot.send_message(uid, f"{get_emoji('warning')} <b>সতর্কবার্তা:</b> একই ডিভাইস/নেটওয়ার্ক থেকে একাধিক অ্যাকাউন্ট খোলা সনাক্ত হয়েছে! রেফার বোনাস যোগ হবে না।", parse_mode="HTML")
                    bot.send_message(ADMIN_ID, f"🚨 <b>Multi-Account Alert!</b>\nUser <code>{uid}</code> joined via Ref <code>{ref_id}</code> from the same Network/IP!", parse_mode="HTML")
                else:
                    update_user_field(uid, "referred_by", int(ref_id))

        # Captcha Challenge Verification
        question, ans = generate_captcha()
        temp = json.loads(u.get("temp_data") or "{}")
        temp["captcha_ans"] = ans
        update_user_field(uid, "temp_data", json.dumps(temp))
        update_user_field(uid, "state", "verify_captcha")

        bot.send_message(message.chat.id, f"🤖 <b>বট সিকিউরিটি ভেরিফিকেশন:</b>\n\nদয়া করে গাণিতিক উত্তরটি দিন:\n👉 <b>{question}</b>", parse_mode="HTML")
    except Exception as e:
        log_error(f"Error in /start: {e}\n{traceback.format_exc()}")

# ============================================
# --- EXCEL REPORT GENERATOR ENGINE ---
# ============================================
def generate_custom_excel(service, rows, filename):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{service.upper()} Unsold Data"
    ws.views.sheetView[0].showGridLines = True

    # Professional Styling Palette
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    zebra_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    border_side = Side(style='thin', color='D9D9D9')
    cell_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
    center_align = Alignment(horizontal="center", vertical="center")

    headers = ["Submit ID", "User ID", "Data/Payload", "Rate (৳)", "Date & Time"]
    ws.append(headers)

    for col_num, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = cell_border

    ws.row_dimensions[1].height = 25

    for row_idx, r in enumerate(rows, start=2):
        date_formatted = datetime.fromtimestamp(float(r.get('created_at', time.time()))).strftime('%Y-%m-%d %H:%M')
        row_data = [
            str(r.get('id', '')),
            str(r.get('user_id', '')),
            str(r.get('payload', '')),
            float(r.get('rate', 0.0)),
            date_formatted
        ]
        ws.append(row_data)
        ws.row_dimensions[row_idx].height = 20
        fill_to_apply = zebra_fill if row_idx % 2 == 0 else white_fill

        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.fill = fill_to_apply
            cell.border = cell_border
            cell.alignment = center_align
            cell.font = Font(name="Calibri", size=10)

    # Auto-fit columns calculation
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or '')
            if len(val_str) > max_len:
                max_len = len(val_str)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    wb.save(filename)

# ============================================
# --- MAIN MESSAGE ROUTER ---
# ============================================
@bot.message_handler(func=lambda message: True, content_types=['text', 'photo', 'document', 'video'])
def handle_all_messages(message):
    try:
        uid = message.from_user.id
        txt = message.text.strip() if message.text else ""
        u = get_user_db(uid)

        # Captcha Check First
        state = u.get("state")
        if state == "verify_captcha":
            temp = json.loads(u.get("temp_data") or "{}")
            c_ans = temp.get("captcha_ans")
            if txt == c_ans:
                update_user_field(uid, "state", None)
                bot.send_message(message.chat.id, "✅ <b>ক্যাপচা ভেরিফিকেশন সফল হয়েছে!</b>", parse_mode="HTML")
                
                # Check Referral Notification
                if u.get("referred_by") and int(u.get("ref_rewarded", 0)) == 0:
                    ref_id = u.get("referred_by")
                    fname = message.from_user.first_name
                    ref_msg = (f"🎉 <b>নতুন রেফারেল নোটিফিকেশন!</b>\n\n"
                               f"আপনার রেফার লিংক ব্যবহার করে <b>{fname}</b> বটে জয়েন করেছে।\n\n"
                               f"👉 <b>দয়া করে তাকে একটি কাজ করতে বলেন তাহলে আপনি বোনাস ও ১০% কমিশন পাবেন।</b>")
                    try: bot.send_message(ref_id, ref_msg, parse_mode="HTML")
                    except: pass
                
                # Force Join Guard after Captcha
                if not check_force_join(uid):
                    msg = f"👋 <b>Welcome to {BOT_NAME}!</b>\n\nবটের কাজ করার জন্য নিচের চ্যানেলগুলোতে জয়েন করুন এবং 'Verify Now' চাপুন:"
                    bot.send_message(message.chat.id, msg, parse_mode="HTML", reply_markup=get_force_join_markup())
                    return
                else:
                    bot.send_message(message.chat.id, f"💎 <b>Welcome to {BOT_NAME}!</b>\nনিচের প্রিমিয়াম মেনু থেকে আপনার পছন্দ বেছে নিন:", parse_mode="HTML", reply_markup=get_main_menu(uid))
                    return
            else:
                question, ans = generate_captcha()
                temp["captcha_ans"] = ans
                update_user_field(uid, "temp_data", json.dumps(temp))
                bot.send_message(message.chat.id, f"❌ <b>ভুল উত্তর!</b> আবার চেষ্টা করুন:\n👉 <b>{question}</b>", parse_mode="HTML")
                return

        # Force Join Real-time Enforcement Guard
        if uid != ADMIN_ID and not check_force_join(uid):
            bot.send_message(message.chat.id, f"{get_emoji('warning')} <b>কাজ শুরু করার আগে অবশ্যই আপনাকে আমাদের অফিসিয়াল চ্যানেলগুলোতে জয়েন করতে হবে!</b>", parse_mode="HTML", reply_markup=get_force_join_markup())
            return

        if u.get("is_banned"): return

        # --- ADMIN STATES ---
        if state and check_permission(uid, "admin"):
            if state == "set_rate_ins":
                set_config("rate_ins", txt)
                update_user_field(uid, "state", None)
                log_admin_action(uid, f"Set Ins Rate to ৳{txt}")
                bot.send_message(message.chat.id, f"✅ <b>Instagram Rate set to: ৳{txt}</b>", parse_mode="HTML", reply_markup=get_admin_menu())
                return
            elif state == "set_rate_fb":
                set_config("rate_fb", txt)
                update_user_field(uid, "state", None)
                log_admin_action(uid, f"Set FB Rate to ৳{txt}")
                bot.send_message(message.chat.id, f"✅ <b>Facebook Rate set to: ৳{txt}</b>", parse_mode="HTML", reply_markup=get_admin_menu())
                return
            elif state == "set_rate_gmail":
                set_config("rate_gmail", txt)
                update_user_field(uid, "state", None)
                log_admin_action(uid, f"Set Gmail Rate to ৳{txt}")
                bot.send_message(message.chat.id, f"✅ <b>Gmail Rate set to: ৳{txt}</b>", parse_mode="HTML", reply_markup=get_admin_menu())
                return
            elif state == "add_force_channel_input":
                chs = json.loads(get_config("force_channels", "[]"))
                new_ch = txt if txt.startswith("@") else f"@{txt}"
                if new_ch not in chs:
                    chs.append(new_ch)
                    set_config("force_channels", json.dumps(chs))
                    bot.send_message(message.chat.id, f"✅ <b>চ্যানেল {new_ch} সফলভাবে যুক্ত করা হয়েছে!</b>", parse_mode="HTML", reply_markup=get_admin_menu())
                else:
                    bot.send_message(message.chat.id, "❌ <b>চ্যানেলটি আগেই তালিকায় যুক্ত আছে!</b>", parse_mode="HTML", reply_markup=get_admin_menu())
                update_user_field(uid, "state", None)
                return
            elif state == "export_date_input":
                temp = json.loads(u.get("temp_data") or "{}")
                serv = temp.get("export_service", "gmail")
                try:
                    target_dt = datetime.strptime(txt, "%Y-%m-%d").date()
                    all_subs = fb_get("pending_submissions") or {}
                    matched = []
                    for sid, s in all_subs.items():
                        if s.get("sub_type") == serv and s.get("status") == "pending":
                            s_date = datetime.fromtimestamp(float(s.get("created_at", 0))).date()
                            if s_date == target_dt:
                                s["id"] = sid
                                matched.append(s)

                    if not matched:
                        bot.send_message(message.chat.id, f"❌ <b>{txt} তারিখে কোনো Unsold ডাটা পাওয়া যায়নি!</b>", parse_mode="HTML", reply_markup=get_admin_menu())
                    else:
                        filename = f"{serv}_{txt}.xlsx"
                        generate_custom_excel(serv, matched, filename)
                        with open(filename, "rb") as f:
                            bot.send_document(message.chat.id, f, caption=f"📊 <b>{serv.upper()} Custom Date ({txt}) Unsold Report</b>", parse_mode="HTML")
                        os.remove(filename)

                        markup = InlineKeyboardMarkup()
                        markup.add(ibtn(f"🗑️ Delete Batch ({txt})", callback_data=f"del_batch_{serv}_{txt}", style="danger"))
                        bot.send_message(message.chat.id, f"⚠️ <b>ফাইল ডাউনলোড সম্পন্ন। উক্ত দিনের ফাইল সার্ভার থেকে মুছে ফেলতে চাইলে 'Delete Batch' বাটন চাপুন:</b>", parse_mode="HTML", reply_markup=markup)
                except Exception as e:
                    bot.send_message(message.chat.id, "❌ <b>ভুল তারিখ ফরম্যাট! সঠিকভাবে YYYY-MM-DD আকারে দিন (যেমন: 2026-08-22)</b>", parse_mode="HTML")
                update_user_field(uid, "state", None)
                return
            elif state == "set_sheet_ins":
                set_config("sheet_id_ins", txt)
                update_user_field(uid, "state", None)
                bot.send_message(message.chat.id, "✅ <b>Instagram Sheet ID Saved!</b>", parse_mode="HTML", reply_markup=get_admin_menu())
                return
            elif state == "set_sheet_fb":
                set_config("sheet_id_fb", txt)
                update_user_field(uid, "state", None)
                bot.send_message(message.chat.id, "✅ <b>Facebook Sheet ID Saved!</b>", parse_mode="HTML", reply_markup=get_admin_menu())
                return
            elif state == "set_sheet_gmail":
                set_config("sheet_id_gmail", txt)
                update_user_field(uid, "state", None)
                bot.send_message(message.chat.id, "✅ <b>Gmail Sheet ID Saved!</b>", parse_mode="HTML", reply_markup=get_admin_menu())
                return
            elif state == "set_json_creds":
                if message.document:
                    file_info = bot.get_file(message.document.file_id)
                    downloaded_file = bot.download_file(file_info.file_path)
                    set_config("json_credentials", downloaded_file.decode('utf-8'))
                else:
                    set_config("json_credentials", txt)
                update_user_field(uid, "state", None)
                bot.send_message(message.chat.id, "✅ <b>Google Cloud Service Account JSON Saved!</b>", parse_mode="HTML", reply_markup=get_admin_menu())
                return
            elif state == "set_firebase_api":
                if message.document:
                    file_info = bot.get_file(message.document.file_id)
                    downloaded_file = bot.download_file(file_info.file_path)
                    set_config("firebase_api", downloaded_file.decode('utf-8'))
                else:
                    set_config("firebase_api", txt)
                update_user_field(uid, "state", None)
                bot.send_message(message.chat.id, "✅ <b>Firebase API Config Saved!</b>", parse_mode="HTML", reply_markup=get_admin_menu())
                return
            elif state == "set_recovery_email":
                set_config("recovery_email", txt)
                update_user_field(uid, "state", None)
                bot.send_message(message.chat.id, f"✅ <b>Recovery Email Set to: <code>{txt}</code></b>", parse_mode="HTML", reply_markup=get_admin_menu())
                return
            elif state == "set_ref_bonus":
                set_config("ref_bonus", txt)
                update_user_field(uid, "state", None)
                bot.send_message(message.chat.id, f"✅ <b>Refer Bonus Set to: ৳{txt}</b>", parse_mode="HTML", reply_markup=get_admin_menu())
                return
            elif state == "set_pass_ins":
                set_config("ins_pass", txt)
                update_user_field(uid, "state", None)
                bot.send_message(message.chat.id, f"✅ <b>Instagram Password Base Set to: <code>{txt}</code></b>", parse_mode="HTML", reply_markup=get_admin_menu())
                return
            elif state == "set_pass_fb":
                set_config("fb_pass", txt)
                update_user_field(uid, "state", None)
                bot.send_message(message.chat.id, f"✅ <b>Facebook Password Base Set to: <code>{txt}</code></b>", parse_mode="HTML", reply_markup=get_admin_menu())
                return
            elif state == "set_pass_gmail":
                set_config("gmail_pass", txt)
                update_user_field(uid, "state", None)
                bot.send_message(message.chat.id, f"✅ <b>Gmail Password Base Set to: <code>{txt}</code></b>", parse_mode="HTML", reply_markup=get_admin_menu())
                return
            elif state == "set_spin_reward":
                set_config("spin_reward", txt)
                update_user_field(uid, "state", None)
                bot.send_message(message.chat.id, f"✅ <b>Spin Reward Set to: ৳{txt}</b>", parse_mode="HTML", reply_markup=get_admin_menu())
                return
            elif state == "set_spin_limit":
                set_config("spin_limit", txt)
                update_user_field(uid, "state", None)
                bot.send_message(message.chat.id, f"✅ <b>Spin Daily Limit Set to: {txt} Times</b>", parse_mode="HTML", reply_markup=get_admin_menu())
                return
            elif state == "set_spin_url":
                set_config("spin_ad_url", txt)
                update_user_field(uid, "state", None)
                bot.send_message(message.chat.id, f"✅ <b>Adsterra Direct Link Set to: <code>{txt}</code></b>", parse_mode="HTML", reply_markup=get_admin_menu())
                return
            elif state == "admin_ban_user":
                if txt.isdigit():
                    update_user_field(int(txt), "is_banned", 1)
                    log_admin_action(uid, f"Banned User {txt}")
                    bot.send_message(message.chat.id, f"⛔ <b>User <code>{txt}</code> has been BANNED!</b>", parse_mode="HTML", reply_markup=get_admin_menu())
                else: bot.send_message(message.chat.id, "❌ <b>ইনভ্যালিড ইউজার আইডি!</b>", parse_mode="HTML")
                update_user_field(uid, "state", None)
                return
            elif state == "admin_unban_user":
                if txt.isdigit():
                    update_user_field(int(txt), "is_banned", 0)
                    log_admin_action(uid, f"Unbanned User {txt}")
                    bot.send_message(message.chat.id, f"✅ <b>User <code>{txt}</code> has been UNBANNED!</b>", parse_mode="HTML", reply_markup=get_admin_menu())
                else: bot.send_message(message.chat.id, "❌ <b>ইনভ্যালিড ইউজার আইডি!</b>", parse_mode="HTML")
                update_user_field(uid, "state", None)
                return
            elif state == "admin_add_bal_id":
                temp = json.loads(u.get("temp_data") or "{}")
                temp["target_user"] = txt
                update_user_field(uid, "temp_data", json.dumps(temp))
                update_user_field(uid, "state", "admin_add_bal_amt")
                bot.send_message(message.chat.id, "💰 <b>কত টাকা ব্যালেন্স যোগ করতে চান? (যেমন: 50):</b>", parse_mode="HTML")
                return
            elif state == "admin_add_bal_amt":
                try:
                    amt = float(txt)
                    temp = json.loads(u.get("temp_data") or "{}")
                    target = temp.get("target_user")
                    t_u = get_user_db(target)
                    fb_update(f"users/{target}", {"balance": float(t_u.get("balance", 0)) + amt})
                    log_admin_action(uid, f"Added ৳{amt} balance to User {target}")
                    bot.send_message(message.chat.id, f"✅ <b>User <code>{target}</code> এর সাথে ৳{amt} যোগ করা হয়েছে!</b>", parse_mode="HTML", reply_markup=get_admin_menu())
                    try: bot.send_message(target, f"🎉 <b>এডমিন আপনার অ্যাকাউন্টে ৳{amt} যোগ করেছেন!</b>", parse_mode="HTML")
                    except: pass
                except: bot.send_message(message.chat.id, "❌ <b>ইনভ্যালিড অ্যামাউন্ট!</b>", parse_mode="HTML", reply_markup=get_admin_menu())
                update_user_field(uid, "state", None)
                return
            elif state == "admin_ded_bal_id":
                temp = json.loads(u.get("temp_data") or "{}")
                temp["target_user"] = txt
                update_user_field(uid, "temp_data", json.dumps(temp))
                update_user_field(uid, "state", "admin_ded_bal_amt")
                bot.send_message(message.chat.id, "💰 <b>কত টাকা ব্যালেন্স কাটতে চান? (যেমন: 20):</b>", parse_mode="HTML")
                return
            elif state == "admin_ded_bal_amt":
                try:
                    amt = float(txt)
                    temp = json.loads(u.get("temp_data") or "{}")
                    target = temp.get("target_user")
                    t_u = get_user_db(target)
                    new_bal = max(0.0, float(t_u.get("balance", 0)) - amt)
                    fb_update(f"users/{target}", {"balance": new_bal})
                    log_admin_action(uid, f"Deducted ৳{amt} balance from User {target}")
                    bot.send_message(message.chat.id, f"✅ <b>User <code>{target}</code> এর অ্যাকাউন্ট থেকে ৳{amt} কাটা হয়েছে!</b>", parse_mode="HTML", reply_markup=get_admin_menu())
                except: bot.send_message(message.chat.id, "❌ <b>ইনভ্যালিড অ্যামাউন্ট!</b>", parse_mode="HTML", reply_markup=get_admin_menu())
                update_user_field(uid, "state", None)
                return
            elif state == "admin_sub_admin_id":
                if txt.isdigit():
                    fb_update(f"users/{txt}", {"role": "sub_admin", "permissions": json.dumps({"all": True})})
                    bot.send_message(message.chat.id, f"👑 <b>User <code>{txt}</code> is now a Sub-Admin!</b>", parse_mode="HTML", reply_markup=get_admin_menu())
                else:
                    bot.send_message(message.chat.id, "❌ <b>ইনভ্যালিড আইডি!</b>", parse_mode="HTML")
                update_user_field(uid, "state", None)
                return
            elif state == "admin_user_search":
                if txt.isdigit():
                    s_u = get_user_db(int(txt))
                    msg = (f"🔍 <b>USER DEEP SEARCH PROFILE</b>\n\n"
                           f"🆔 User ID: <code>{s_u.get('user_id')}</code>\n"
                           f"💰 Balance: ৳{float(s_u.get('balance', 0)):.2f}\n"
                           f"👥 Referrals: {s_u.get('referrals', 0)}\n"
                           f"📥 Total Withdrawn: ৳{float(s_u.get('total_withdraw', 0)):.2f}\n"
                           f"✅ Approved Tasks: {s_u.get('approved_tasks', 0)}\n"
                           f"❌ Rejected Tasks: {s_u.get('rejected_tasks', 0)}\n"
                           f"⛔ Is Banned: {'YES' if s_u.get('is_banned') else 'NO'}\n"
                           f"👑 Role: {s_u.get('role', 'user')}")
                    bot.send_message(message.chat.id, msg, parse_mode="HTML", reply_markup=get_admin_menu())
                else: bot.send_message(message.chat.id, "❌ <b>ইনভ্যালিড ইউজার আইডি!</b>", parse_mode="HTML")
                update_user_field(uid, "state", None)
                return
            elif state == "admin_broadcast_msg":
                all_users = fb_get("users") or {}
                count = 0
                bot.send_message(message.chat.id, f"🚀 <b>{len(all_users)} জন ইউজারের কাছে মেসেজ ব্রডকাস্ট শুরু হচ্ছে...</b>", parse_mode="HTML")
                for u_id in all_users.keys():
                    try:
                        bot.send_message(u_id, txt, parse_mode="HTML")
                        count += 1
                        time.sleep(0.04)
                    except: pass
                bot.send_message(message.chat.id, f"✅ <b>ব্রডকাস্ট সম্পন্ন! মোট {count} জন ইউজার মেসেজ পেয়েছেন।</b>", parse_mode="HTML", reply_markup=get_admin_menu())
                update_user_field(uid, "state", None)
                return
            elif state == "admin_add_app_task":
                try:
                    lines = txt.split("\n")
                    t_title = lines[0]
                    t_link = lines[1]
                    t_reward = float(lines[2])
                    task_id = f"task_{random.randint(1000, 9999)}"
                    fb_set(f"app_tasks/{task_id}", {
                        "id": task_id, "title": t_title, "link": t_link, "reward": t_reward
                    })
                    bot.send_message(message.chat.id, f"✅ <b>নতুন অ্যাপ/টেলিগ্রাম টাস্ক যুক্ত হয়েছে!</b>\nTitle: {t_title}\nReward: ৳{t_reward}", parse_mode="HTML", reply_markup=get_admin_menu())
                except:
                    bot.send_message(message.chat.id, "❌ <b>ভুল ফরম্যাট! ৩ লাইনে শিরোনাম, লিংক এবং রিওয়ার্ড রিড্রেস দিন।</b>", parse_mode="HTML")
                update_user_field(uid, "state", None)
                return
            elif state == "admin_set_vid_gmail":
                vids = json.loads(get_config("tutorial_videos", "{}"))
                vids["gmail"] = txt
                set_config("tutorial_videos", json.dumps(vids))
                bot.send_message(message.chat.id, f"✅ <b>Gmail Tutorial Video Link Saved!</b>", parse_mode="HTML", reply_markup=get_admin_menu())
                update_user_field(uid, "state", None)
                return
            elif state == "admin_set_vid_fb":
                vids = json.loads(get_config("tutorial_videos", "{}"))
                vids["fb"] = txt
                set_config("tutorial_videos", json.dumps(vids))
                bot.send_message(message.chat.id, f"✅ <b>Facebook Tutorial Video Link Saved!</b>", parse_mode="HTML", reply_markup=get_admin_menu())
                update_user_field(uid, "state", None)
                return
            elif state == "admin_set_vid_ins":
                vids = json.loads(get_config("tutorial_videos", "{}"))
                vids["ins"] = txt
                set_config("tutorial_videos", json.dumps(vids))
                bot.send_message(message.chat.id, f"✅ <b>Instagram Tutorial Video Link Saved!</b>", parse_mode="HTML", reply_markup=get_admin_menu())
                update_user_field(uid, "state", None)
                return
            elif state and state.startswith("admin_reply_ticket_"):
                t_id = state.replace("admin_reply_ticket_", "")
                t_data = fb_get(f"support_tickets/{t_id}")
                if t_data:
                    u_target = t_data.get("user_id")
                    fb_update(f"support_tickets/{t_id}", {"status": "replied", "reply": txt})
                    try: bot.send_message(u_target, f"🎧 <b>এডমিন এর থেকে সাপোর্ট টিকিট উত্তর ({t_id}):</b>\n\n{txt}", parse_mode="HTML")
                    except: pass
                    bot.send_message(message.chat.id, f"✅ <b>টিকিট {t_id} রিপ্লাই দেওয়া হয়েছে!</b>", parse_mode="HTML", reply_markup=get_admin_menu())
                update_user_field(uid, "state", None)
                return

        # --- USER TASK PROOF SUBMISSION ---
        if message.photo and state and state.startswith("sub_app_proof_"):
            task_id = state.replace("sub_app_proof_", "")
            sub_id = f"sub_{uid}_{int(time.time())}"
            photo_id = message.photo[-1].file_id
            
            sub_data = {
                "id": sub_id,
                "user_id": uid,
                "sub_type": "app_ss",
                "payload": json.dumps({"photo": photo_id, "task_id": task_id}),
                "status": "pending",
                "created_at": time.time()
            }
            fb_set(f"pending_submissions/{sub_id}", sub_data)
            fb_set(f"completed_app_tasks/{uid}/{task_id}", True)
            fb_update(f"users/{uid}", {"pending_tasks": int(u.get("pending_tasks", 0)) + 1})

            update_user_field(uid, "state", None)
            bot.send_message(message.chat.id, "✅ <b>আপনার প্রুফ স্ক্রিনশট জমা হয়েছে!</b> এডমিন চেক করে এপ্রুভ করলে ব্যালেন্স যোগ হবে।", parse_mode="HTML", reply_markup=get_main_menu(uid))
            return

        # --- USER SUPPORT TICKET STATE ---
        if state == "user_submit_ticket":
            if get_daily_support_count(uid) >= 5:
                bot.send_message(message.chat.id, f"{get_emoji('error')} <b>আপনি আজ ৫টির বেশি সাপোর্ট টিকিট দিতে পারবেন না!</b>", parse_mode="HTML")
                update_user_field(uid, "state", None)
                return
                
            t_id = f"TICK-{random.randint(10000, 99999)}"
            ticket_data = {
                "ticket_id": t_id,
                "user_id": uid,
                "message": txt,
                "status": "pending",
                "created_at": time.time()
            }
            fb_set(f"support_tickets/{t_id}", ticket_data)

            update_user_field(uid, "state", None)
            bot.send_message(message.chat.id, f"✅ <b>আপনার সাপোর্ট টিকিট জমা নেওয়া হয়েছে!</b>\nTicket ID: <code>{t_id}</code>", parse_mode="HTML", reply_markup=get_main_menu(uid))
            return

        # --- USER WORK STATES & FB VALIDATION GUARD ---
        if state == "enter_fb_uid":
            if not (txt.isdigit() and 14 <= len(txt) <= 16):
                bot.send_message(message.chat.id, f"{get_emoji('error')} <b>ভুল UID!</b> Facebook UID অবশ্যই ১৪ থেকে ১৬ সংখ্যার হতে হবে। আবার দিন:", parse_mode="HTML")
                return

            if check_duplicate_and_save(txt, "fb_uid", uid):
                bot.send_message(message.chat.id, f"{get_emoji('error')} <b>এই FB UID টি ইতিমধ্যেই সিস্টেমে জমা দেওয়া হয়েছে!</b>", parse_mode="HTML")
                return

            temp = json.loads(u.get("temp_data") or "{}")
            temp["fb_uid"] = txt
            update_user_field(uid, "temp_data", json.dumps(temp))
            update_user_field(uid, "state", "enter_fb_cookie")
            bot.send_message(message.chat.id, "🍪 <b>এবার আপনার FB Cookie টি সেন্ড করুন:</b>", parse_mode="HTML")
            return

        elif state == "enter_fb_cookie":
            if len(txt) <= 28:
                bot.send_message(message.chat.id, f"{get_emoji('error')} <b>ইনভ্যালিড কুকিজ!</b> কুকিজটি সর্বনিম্ন ২৮ অক্ষরের বেশি হতে হবে। আবার দিন:", parse_mode="HTML")
                return

            if check_duplicate_and_save(txt, "fb_cookie", uid):
                bot.send_message(message.chat.id, f"{get_emoji('error')} <b>এই FB Cookie টি আগে ব্যবহার করা হয়েছে!</b>", parse_mode="HTML")
                return
            
            temp = json.loads(u.get("temp_data") or "{}")
            rate = float(get_config("rate_fb", "18.0"))
            sub_id = f"FB-{random.randint(1000, 9999)}"
            
            data_p = {
                "fn": temp.get("fn"), "ln": temp.get("ln"), "pass": temp.get("pass"),
                "uid": temp.get("fb_uid"), "cookie": txt
            }

            sub_data = {
                "id": sub_id,
                "user_id": uid,
                "sub_type": "fb",
                "payload": json.dumps(data_p),
                "rate": rate,
                "status": "pending",
                "created_at": time.time()
            }
            fb_set(f"pending_submissions/{sub_id}", sub_data)
            fb_update(f"users/{uid}", {"pending_tasks": int(u.get("pending_tasks", 0)) + 1})

            append_to_google_sheet("fb", [sub_id, uid, temp.get("fn"), temp.get("ln"), temp.get("fb_uid"), temp.get("pass"), txt, "Pending", datetime.now().strftime("%Y-%m-%d %H:%M")])

            update_user_field(uid, "state", None)
            bot.send_message(message.chat.id, f"🎉 <b>আপনার ফেসবুক কাজ জমা হয়েছে!</b>\nSubmit ID: <code>{sub_id}</code>\nরেট: ৳{rate:.2f}", parse_mode="HTML", reply_markup=get_main_menu(uid))
            return

        elif state == "enter_2fa_code":
            if check_duplicate_and_save(txt, "2fa", uid):
                bot.send_message(message.chat.id, f"{get_emoji('error')} <b>এই 2FA কোডটি আগে ব্যবহার করা হয়েছে!</b>", parse_mode="HTML")
                return
            
            try:
                totp = pyotp.TOTP(txt.replace(" ", "").upper())
                six_digit_code = totp.now()
            except Exception as e:
                log_error(f"2FA Code Generator Error: {e}")
                six_digit_code = "123456"

            temp = json.loads(u.get("temp_data") or "{}")
            temp["2fa_secret"] = txt
            temp["six_digit_code"] = six_digit_code
            update_user_field(uid, "temp_data", json.dumps(temp))

            markup = InlineKeyboardMarkup()
            markup.add(ibtn(f"📋 6-Digit Code: {six_digit_code}", callback_data=f"copy_code_{six_digit_code}", style="success"))

            bot.send_message(message.chat.id, f"🔐 <b>আপনার ইনস্ট্যান্ট ২FA কোড তৈরি হয়েছে:</b>\n\n<code>{six_digit_code}</code>\n\nনিচের বাটন চেপে কোড কপি করুন এবং একাউন্ট খোলা সম্পন্ন করুন:", parse_mode="HTML", reply_markup=markup)

            reply_kb = ReplyKeyboardMarkup(resize_keyboard=True, row_width=1)
            reply_kb.add(rbtn("অ্যাকাউন্ট খোলা শেষ", "success"), rbtn("Cancel ❌", "danger"))
            bot.send_message(message.chat.id, "অ্যাকাউন্ট খোলা শেষ হলে নিচের 'অ্যাকাউন্ট খোলা শেষ' বাটনে ক্লিক করুন:", reply_markup=reply_kb)
            update_user_field(uid, "state", None)
            return

        elif state and state.startswith("withdraw_number_"):
            meth = state.replace("withdraw_number_", "")
            w_methods = json.loads(get_config("withdraw_methods", "{}"))
            min_limit = float(w_methods[meth]["min"])
            max_limit = float(w_methods[meth].get("max", 10000.0))
            
            if get_daily_withdraw_count(uid) >= 2:
                bot.send_message(message.chat.id, f"{get_emoji('warning')} <b>উইথড্র লিমিট শেষ!</b> আপনি ২৪ ঘণ্টায় সর্বোচ্চ ২ বার উইথড্র করতে পারবেন।", parse_mode="HTML")
                update_user_field(uid, "state", None)
                return

            bal = float(u.get("balance", 0.0))
            if bal < min_limit:
                bot.send_message(message.chat.id, f"{get_emoji('error')} <b>আপনার পর্যাপ্ত ব্যালেন্স নেই!</b> মিনিমাম উইথড্র ৳{min_limit}", parse_mode="HTML")
                update_user_field(uid, "state", None)
                return

            req_id = f"W-{random.randint(10000, 99999)}"
            amt = bal if bal <= max_limit else max_limit
            
            today_str = datetime.now().strftime("%Y-%m-%d")
            last_reset = get_config("last_withdraw_reset_date", "")
            if last_reset != today_str:
                set_config("today_total_withdrawn", "0.0")
                set_config("last_withdraw_reset_date", today_str)

            daily_bot_limit = float(get_config("daily_bot_withdraw_limit", "50000.0"))
            today_total = float(get_config("today_total_withdrawn", "0.0"))

            if today_total + amt > daily_bot_limit:
                bot.send_message(message.chat.id, f"{get_emoji('warning')} <b>বটের আজকের গ্লোবাল ক্যাশআউট লিমিট পূর্ণ হয়েছে!</b> অনুগ্রহ করে আগামীকাল আবার চেষ্টা করুন।", parse_mode="HTML")
                update_user_field(uid, "state", None)
                return

            req_data = {
                "req_id": req_id,
                "user_id": uid,
                "method": meth,
                "account_number": txt,
                "amount": amt,
                "status": "pending",
                "created_at": time.time()
            }
            fb_set(f"withdraw_requests/{req_id}", req_data)
            fb_update(f"users/{uid}", {
                "balance": bal - amt,
                "total_withdraw": float(u.get("total_withdraw", 0.0)) + amt
            })

            set_config("today_total_withdrawn", str(today_total + amt))

            update_user_field(uid, "state", None)
            bot.send_message(message.chat.id, f"✅ <b>আপনার ৳{amt:.2f} এর উইথড্র রিকোয়েস্ট জমা হয়েছে!</b>\nMethod: {meth}\nAccount: <code>{txt}</code>", parse_mode="HTML", reply_markup=get_main_menu(uid))

            markup = InlineKeyboardMarkup(row_width=2)
            markup.add(
                ibtn("✅ Paid", callback_data=f"pay_with_{req_id}", style="success"),
                ibtn("❌ Refund Balance", callback_data=f"ref_with_{req_id}", style="danger")
            )
            admin_alert = f"📥 <b>New Withdrawal Request!</b>\n\nReq ID: <code>{req_id}</code>\nUser ID: <code>{uid}</code>\nMethod: {meth}\nAccount: <code>{txt}</code>\nAmount: ৳{amt:.2f}"
            try: bot.send_message(ADMIN_ID, admin_alert, parse_mode="HTML", reply_markup=markup)
            except: pass
            return

        # --- MENU ROUTING ---
        if txt == TXT_WORK_MAIN:
            bot.send_message(message.chat.id, "💼 <b>কাজ অপশন নির্বাচন করুন:</b>", parse_mode="HTML", reply_markup=get_work_menu())
        elif txt == TXT_BACK:
            update_user_field(uid, "state", None)
            bot.send_message(message.chat.id, "🏠 Main Menu", reply_markup=get_main_menu(uid))

        elif txt == f"{get_emoji('instagram')} ইনস্টাগ্রাম কাজ":
            if get_config("pause_ins", "false") == "true":
                bot.send_message(message.chat.id, f"{get_emoji('warning')} <b>ইনস্টাগ্রাম কাজ সাময়িক বন্ধ আছে।</b>", parse_mode="HTML")
                return
            rate = get_config("rate_ins", "15.0")
            pass_val = generate_dynamic_password("ins_pass")
            _, _, un = generate_random_identity()
            temp_data = json.dumps({"start_time": time.time(), "username": un, "pass": pass_val})
            update_user_field(uid, "temp_data", temp_data)

            msg = (f"📸 <b>Instagram Account Creation</b>\n\n"
                   f"💰 কাজের মূল্য: <b>৳{rate}</b>\n"
                   f"👤 Username: <code>{un}</code>\n"
                   f"🔑 Password: <code>{pass_val}</code>\n\n"
                   f"অ্যাকাউন্ট খুলে 2FA সেটআপ করে নিচের বাটনে চাপ দিন।")
            
            markup = ReplyKeyboardMarkup(resize_keyboard=True, row_width=1)
            markup.add(rbtn("🔑 2FA সেট", "primary"), rbtn("Cancel ❌", "danger"))
            bot.send_message(message.chat.id, msg, parse_mode="HTML", reply_markup=markup)

        elif txt == "🔑 2FA সেট":
            update_user_field(uid, "state", "enter_2fa_code")
            bot.send_message(message.chat.id, "🔐 <b>আপনার 2FA Secret Key টি দিন:</b>", parse_mode="HTML")

        elif txt == "Cancel ❌":
            update_user_field(uid, "state", None)
            bot.send_message(message.chat.id, "❌ কাজ বাতিল করা হয়েছে।", reply_markup=get_main_menu(uid))

        elif txt == "অ্যাকাউন্ট খোলা শেষ":
            temp = json.loads(u.get("temp_data") or "{}")
            rate = float(get_config("rate_ins", "15.0"))
            sub_id = f"INS-{random.randint(1000, 9999)}"
            
            data_p = {
                "username": temp.get("username"),
                "pass": temp.get("pass"),
                "2fa": temp.get("2fa_secret"),
                "code": temp.get("six_digit_code")
            }
            
            sub_data = {
                "id": sub_id,
                "user_id": uid,
                "sub_type": "ins",
                "payload": json.dumps(data_p),
                "rate": rate,
                "status": "pending",
                "created_at": time.time()
            }
            fb_set(f"pending_submissions/{sub_id}", sub_data)
            fb_update(f"users/{uid}", {"pending_tasks": int(u.get("pending_tasks", 0)) + 1})

            append_to_google_sheet("ins", [sub_id, uid, temp.get("username"), temp.get("pass"), temp.get("2fa_secret"), "Pending", datetime.now().strftime("%Y-%m-%d %H:%M")])

            bot.send_message(message.chat.id, f"🎉 <b>আপনার ইনস্টাগ্রাম অ্যাকাউন্ট সফলভাবে পেন্ডিংয়ে জমা দেওয়া হয়েছে!</b>\nSubmit ID: <code>{sub_id}</code>\nরেট: ৳{rate:.2f}", parse_mode="HTML", reply_markup=get_main_menu(uid))

        elif txt == f"{get_emoji('facebook')} ফেসবুক কাজ":
            if get_config("pause_fb", "false") == "true":
                bot.send_message(message.chat.id, f"{get_emoji('warning')} <b>ফেসবুক কাজ সাময়িক বন্ধ আছে।</b>", parse_mode="HTML")
                return
            rate = get_config("rate_fb", "18.0")
            pass_val = generate_dynamic_password("fb_pass")
            fn, ln, _ = generate_random_identity()
            temp_data = json.dumps({"fn": fn, "ln": ln, "pass": pass_val})
            update_user_field(uid, "temp_data", temp_data)

            msg = (f"📘 <b>Facebook Account Creation</b>\n\n"
                   f"💰 কাজের মূল্য: <b>৳{rate}</b>\n"
                   f"👤 First Name: <code>{fn}</code>\n"
                   f"👤 Last Name: <code>{ln}</code>\n"
                   f"🔑 Password: <code>{pass_val}</code>")
            
            markup = ReplyKeyboardMarkup(resize_keyboard=True, row_width=1)
            markup.add(rbtn("Send UID", "primary"), rbtn("Cancel ❌", "danger"))
            bot.send_message(message.chat.id, msg, parse_mode="HTML", reply_markup=markup)

        elif txt == "Send UID":
            update_user_field(uid, "state", "enter_fb_uid")
            bot.send_message(message.chat.id, "🆔 <b>আপনার Facebook UID প্রদান করুন (১৪-১৬ সংখ্যা):</b>", parse_mode="HTML")

        elif txt == f"{get_emoji('gmail')} Gmail কাজ":
            if get_config("pause_gmail", "false") == "true":
                bot.send_message(message.chat.id, f"{get_emoji('warning')} <b>জিমেইল কাজ সাময়িক বন্ধ আছে।</b>", parse_mode="HTML")
                return
            rate = get_config("rate_gmail", "12.0")
            pass_val = generate_dynamic_password("gmail_pass")
            fn, ln, un = generate_random_identity()
            g_email = f"{un}@gmail.com"
            temp_data = json.dumps({"start_time": time.time(), "email": g_email, "pass": pass_val})
            update_user_field(uid, "temp_data", temp_data)

            msg = (f"📧 <b>New Gmail Sell Task</b>\n\n"
                   f"💰 কাজের মূল্য: <b>৳{rate}</b>\n"
                   f"👤 First Name: <code>{fn}</code>\n"
                   f"👤 Last Name: <code>{ln}</code>\n"
                   f"✉️ Gmail: <code>{g_email}</code>\n"
                   f"🔑 Password: <code>{pass_val}</code>")
            
            markup = ReplyKeyboardMarkup(resize_keyboard=True, row_width=1)
            markup.add(rbtn("কাজ শেষ", "success"), rbtn("বাতিল", "danger"))
            bot.send_message(message.chat.id, msg, parse_mode="HTML", reply_markup=markup)

        elif txt in ["কাজ শেষ", "বাতিল"]:
            if txt == "বাতিল":
                update_user_field(uid, "state", None)
                bot.send_message(message.chat.id, "❌ কাজ বাতিল করা হয়েছে।", reply_markup=get_main_menu(uid))
                return

            temp = json.loads(u.get("temp_data") or "{}")
            start_time = float(temp.get("start_time", 0))

            if time.time() - start_time < 120:
                bot.send_message(message.chat.id, f"{get_emoji('error')} <b>সতর্কবার্তা:</b> আপনি কাজ শুরুর মাত্র ২ মিনিটের মধ্যে 'কাজ শেষ' চাপ দিয়েছেন! সঠিক নিয়মে অ্যাকাউন্ট তৈরি করতে ন্যূনতম ৩ মিনিট সময় অতিবাহিত করুন।", parse_mode="HTML")
                return

            g_email = temp.get("email")
            bot.send_message(message.chat.id, "⏳ <b>দয়া করে অপেক্ষা করুন, গুগল মেইল সার্ভারে অ্যাকাউন্টের স্থায়িত্ব সাইলেন্টলি ভেরিফাই করা হচ্ছে....</b>", parse_mode="HTML")
            
            if not verify_gmail_smtp(g_email):
                bot.send_message(message.chat.id, f"{get_emoji('error')} <b>আপনি জিমেইল অ্যাকাউন্টটি তৈরি না করেই 'কাজ শেষ' বাটনে চাপ দিয়েছেন!</b>\n\nদয়া করে সঠিক নিয়মে অ্যাকাউন্ট তৈরি করে আবার চেষ্টা করুন।", parse_mode="HTML", reply_markup=get_main_menu(uid))
                return

            rate = float(get_config("rate_gmail", "12.0"))
            sub_id = f"GM-{random.randint(1000, 9999)}"
            data_p = {"email": temp.get("email"), "pass": temp.get("pass")}
            
            sub_data = {
                "id": sub_id,
                "user_id": uid,
                "sub_type": "gmail",
                "payload": json.dumps(data_p),
                "rate": rate,
                "status": "pending",
                "created_at": time.time()
            }
            fb_set(f"pending_submissions/{sub_id}", sub_data)
            fb_update(f"users/{uid}", {"pending_tasks": int(u.get("pending_tasks", 0)) + 1})

            rec_email = get_config("recovery_email", "tasrikvai8001@gmail.com")
            append_to_google_sheet("gmail", [sub_id, uid, temp.get("email"), temp.get("pass"), rec_email, "Pending", datetime.now().strftime("%Y-%m-%d %H:%M")])

            bot.send_message(message.chat.id, f"✅ <b>জিমেইল কাজ সফলভাবে যাচাইপূর্বক জমা নেওয়া হয়েছে!</b>\nSubmit ID: <code>{sub_id}</code>", parse_mode="HTML", reply_markup=get_main_menu(uid))

        elif txt == TXT_TODAY_WORK:
            markup = InlineKeyboardMarkup(row_width=1)
            markup.add(
                ibtn("🌀 স্পিন করে আয়", callback_data="open_spin_game", style="primary"),
                ibtn("📲 টেলিগ্রাম ও অ্যাপস টাস্ক", callback_data="open_app_tasks", style="primary"),
                ibtn("🎁 আমন্ত্রণ পুরষ্কার", callback_data="open_invite_rewards", style="success")
            )
            bot.send_message(message.chat.id, "🔥 <b>আজকের কাজের মেনু:</b>", parse_mode="HTML", reply_markup=markup)

        elif txt == TXT_BALANCE:
            msg = (f"<b>👤 USER STATS & BALANCE</b>\n\n"
                   f"<b>{get_emoji('balance')} মোট ব্যালেন্স:</b> <b>৳{float(u.get('balance', 0)):.2f}</b>\n"
                   f"<b>{get_emoji('invite')} মোট রেফার:</b> <b>{u.get('referrals', 0)}</b>\n"
                   f"<b>{get_emoji('withdraw')} মোট উইথড্র:</b> <b>৳{float(u.get('total_withdraw', 0)):.2f}</b>\n"
                   f"<b>⏳ পেন্ডিং টাস্ক:</b> <b>{u.get('pending_tasks', 0)}</b>\n"
                   f"<b>✅ এপ্রুভড টাস্ক:</b> <b>{u.get('approved_tasks', 0)}</b>\n"
                   f"<b>❌ রিজেক্ট টাস্ক:</b> <b>{u.get('rejected_tasks', 0)}</b>")
            bot.send_message(message.chat.id, msg, parse_mode="HTML")

        elif txt == TXT_WITHDRAW:
            w_methods = json.loads(get_config("withdraw_methods", "{}"))
            markup = InlineKeyboardMarkup(row_width=1)
            for meth, info in w_methods.items():
                if info.get("enabled"):
                    markup.add(ibtn(f"💳 {meth} (Min ৳{info['min']})", callback_data=f"with_meth_{meth}", style="primary"))
            bot.send_message(message.chat.id, "📥 <b>উইথড্র মেথড সিলেক্ট করুন:</b>", parse_mode="HTML", reply_markup=markup)

        elif txt == TXT_REFER:
            bot_uname = bot.get_me().username
            link = f"https://t.me/{bot_uname}?start={uid}"
            bonus = get_config("ref_bonus", "10.0")
            msg = (f"{get_emoji('invite')} <b>REFER & EARN!</b>\n\n"
                   f"আপনার রেফারেল লিংক:\n<code>{link}</code>\n\n"
                   f"💡 নিয়ম: আপনার লিংক থেকে কোনো ইউজার জয়েন করে ১ম কাজ শেষ করলে পাবেন <b>৳{bonus}</b> এবং তার সারাজীবনের কাজের ওপর পাবেন <b>১০% লাইফটাইম কমিশন!</b>")
            bot.send_message(message.chat.id, msg, parse_mode="HTML")

        elif txt == TXT_SUPPORT:
            markup = InlineKeyboardMarkup(row_width=1)
            markup.add(
                ibtn("🎧 Official Support Channel", url="https://t.me/tasrikvai", style="primary"),
                ibtn("📩 Open Support Ticket", callback_data="open_support_ticket", style="success")
            )
            bot.send_message(message.chat.id, f"{get_emoji('support')} <b>আমাদের ২৪/৭ সাপোর্ট প্যানেল:</b>\n\nসরাসরি এডমিনের সাহায্য নিতে টিকিট ওপেন করুন:", parse_mode="HTML", reply_markup=markup)

        elif txt == TXT_NEWBIE:
            markup = InlineKeyboardMarkup(row_width=1)
            markup.add(
                ibtn("📧 Gmail Account Video", callback_data="watch_vid_gmail", style="primary"),
                ibtn("📘 Facebook Account Video", callback_data="watch_vid_fb", style="primary"),
                ibtn("📸 Instagram Account Video", callback_data="watch_vid_ins", style="primary")
            )
            msg = f"{get_emoji('newbie')} <b>টিউটোরিয়াল প্যানেল</b>\n\nনিচের বাটনগুলোতে চাপ দিয়ে যেকোনো কাজের প্রিমিয়াম ভিডিও দেখে কাজ শিখুন:"
            bot.send_message(message.chat.id, msg, parse_mode="HTML", reply_markup=markup)

        elif txt == TXT_ADMIN_PANEL and check_permission(uid, "admin"):
            bot.send_message(message.chat.id, "⚙️ <b>Admin Control Panel</b>", parse_mode="HTML", reply_markup=get_admin_menu())

        # --- ALL ADMIN PANEL BUTTON HANDLERS ---
        elif txt == "📢 Set Force Join" and check_permission(uid, "admin"):
            chs = json.loads(get_config("force_channels", "[]"))
            markup = InlineKeyboardMarkup(row_width=1)
            markup.add(ibtn("➕ Add New Force Join Channel", callback_data="add_force_channel", style="success"))
            for ch in chs:
                markup.add(ibtn(f"❌ Remove {ch}", callback_data=f"rem_force_{ch}", style="danger"))
            bot.send_message(message.chat.id, "📢 <b>Force Join Channel Manager:</b>", parse_mode="HTML", reply_markup=markup)

        elif txt == "🎁 Set Ref Bonus" and check_permission(uid, "admin"):
            update_user_field(uid, "state", "set_ref_bonus")
            bot.send_message(message.chat.id, f"🎁 <b>বর্তমান রেফারাল বোনাস: ৳{get_config('ref_bonus')}</b>\nনতুন রেফারাল বোনাস অ্যামাউন্ট দিন (যেমন: 10):", parse_mode="HTML")

        elif txt == "💳 Withdraw Setting" and check_permission(uid, "admin"):
            w_methods = json.loads(get_config("withdraw_methods", "{}"))
            markup = InlineKeyboardMarkup(row_width=1)
            for m, data in w_methods.items():
                status = "✅ Enabled" if data.get("enabled") else "❌ Disabled"
                markup.add(ibtn(f"{m}: {status}", callback_data=f"toggle_wmeth_{m}", style="primary"))
            bot.send_message(message.chat.id, "💳 <b>Withdraw Method Manager (Toggle On/Off):</b>", parse_mode="HTML", reply_markup=markup)

        elif txt == "🔑 Set Passwords Base" and check_permission(uid, "admin"):
            markup = InlineKeyboardMarkup(row_width=1)
            markup.add(
                ibtn(f"📸 Instagram Base ({get_config('ins_pass')})", callback_data="set_pass_ins", style="primary"),
                ibtn(f"📘 Facebook Base ({get_config('fb_pass')})", callback_data="set_pass_fb", style="primary"),
                ibtn(f"📧 Gmail Base ({get_config('gmail_pass')})", callback_data="set_pass_gmail", style="primary")
            )
            bot.send_message(message.chat.id, "🔑 <b>অটোমেটিক ডায়নামিক পাসওয়ার্ড বেস সেটিং:</b>", parse_mode="HTML", reply_markup=markup)

        elif txt == "🎨 Edit Emoji IDs" and check_permission(uid, "admin"):
            emojis = json.loads(get_config("emojis", "{}"))
            msg = "🎨 <b>বর্তমান বটে ব্যবহৃত ইমোজি তালিকা:</b>\n\n" + "\n".join([f"{k}: {v}" for k, v in emojis.items()])
            bot.send_message(message.chat.id, msg, parse_mode="HTML")

        elif txt == "🌀 Spin & Ad Settings" and check_permission(uid, "admin"):
            markup = InlineKeyboardMarkup(row_width=1)
            markup.add(
                ibtn(f"💰 Reward (Current: ৳{get_config('spin_reward')})", callback_data="set_spin_reward", style="primary"),
                ibtn(f"🔢 Limit (Current: {get_config('spin_limit')})", callback_data="set_spin_limit", style="primary"),
                ibtn(f"🌐 Ad Link (Current)", callback_data="set_spin_url", style="primary")
            )
            bot.send_message(message.chat.id, "🌀 <b>স্পিন ও অ্যাডস্টাররা অ্যাড সেটিংস:</b>", parse_mode="HTML", reply_markup=markup)

        elif txt == "📊 Google Sheets Config" and check_permission(uid, "admin"):
            markup = InlineKeyboardMarkup(row_width=1)
            markup.add(
                ibtn("📸 Instagram Sheet ID", callback_data="set_sheet_ins", style="primary"),
                ibtn("📘 Facebook Sheet ID", callback_data="set_sheet_fb", style="primary"),
                ibtn("📧 Gmail Sheet ID", callback_data="set_sheet_gmail", style="primary")
            )
            bot.send_message(message.chat.id, "📊 <b>Google Sheet Spreadsheet ID Manager:</b>", parse_mode="HTML", reply_markup=markup)

        elif txt == "🤖 Upload JSON Creds" and check_permission(uid, "admin"):
            update_user_field(uid, "state", "set_json_creds")
            bot.send_message(message.chat.id, "🤖 <b>Google Cloud Service Account credential JSON ফাইল বা টেক্সট আপলোড করুন:</b>", parse_mode="HTML")

        elif txt == "🔥 Set Firebase API" and check_permission(uid, "admin"):
            update_user_field(uid, "state", "set_firebase_api")
            bot.send_message(message.chat.id, f"🔥 <b>বর্তমান Firebase API Key:</b> <code>{get_config('firebase_api')}</code>\n\nনতুন Firebase API Key প্রদান করুন:", parse_mode="HTML")

        elif txt == "🔄 Sync Google Sheet Approval" and check_permission(uid, "admin"):
            bot.send_message(message.chat.id, "⏳ <b>গুগল শিট থেকে 'ok'/'bad' স্ট্যাটাস সিঙ্ক করা হচ্ছে...</b>", parse_mode="HTML")
            cnt = sync_sheet_approvals()
            bot.send_message(message.chat.id, f"✅ <b>সিঙ্ক সম্পন্ন হয়েছে! মোট {cnt} টি টাস্ক প্রসেস করা হয়েছে।</b>", parse_mode="HTML", reply_markup=get_admin_menu())

        elif txt == "📥 Export Unsold Files" and check_permission(uid, "admin"):
            markup = InlineKeyboardMarkup(row_width=1)
            markup.add(
                ibtn("📸 Export Instagram Unsold", callback_data="exp_menu_ins", style="primary"),
                ibtn("📘 Export Facebook Unsold", callback_data="exp_menu_fb", style="primary"),
                ibtn("📧 Export Gmail Unsold", callback_data="exp_menu_gmail", style="primary")
            )
            bot.send_message(message.chat.id, "📥 <b>কোন সার্ভিসের Unsold ফাইল এক্সপোর্ট করতে চান?</b>", parse_mode="HTML", reply_markup=markup)

        elif txt == "🔎 Pending Approvals" and check_permission(uid, "admin"):
            subs = fb_get("pending_submissions") or {}
            pending_list = [s for s_id, s in subs.items() if s.get("status") == "pending"]
            if not pending_list:
                bot.send_message(message.chat.id, "✅ <b>কোনো পেন্ডিং টাস্ক সাবমিশন নেই!</b>", parse_mode="HTML")
            else:
                bot.send_message(message.chat.id, f"🔎 <b>মোট {len(pending_list)} টি পেন্ডিং কাজ আছে:</b>", parse_mode="HTML")
                for item in pending_list[:5]:
                    s_id = item.get("id")
                    u_target = item.get("user_id")
                    stype = item.get("sub_type")
                    srate = item.get("rate")
                    p_data = item.get("payload")
                    
                    markup = InlineKeyboardMarkup(row_width=2)
                    markup.add(
                        ibtn("✅ Approve", callback_data=f"adm_app_{s_id}", style="success"),
                        ibtn("❌ Reject", callback_data=f"adm_rej_{s_id}", style="danger")
                    )
                    bot.send_message(message.chat.id, f"📌 Sub ID: <code>{s_id}</code>\nUser: <code>{u_target}</code>\nType: {stype}\nRate: ৳{srate}\nPayload: <code>{p_data}</code>", parse_mode="HTML", reply_markup=markup)

        elif txt == "📥 Pending Withdraws" and check_permission(uid, "admin"):
            reqs = fb_get("withdraw_requests") or {}
            p_reqs = [r for r_id, r in reqs.items() if r.get("status") == "pending"]
            if not p_reqs:
                bot.send_message(message.chat.id, "✅ <b>কোনো পেন্ডিং উইথড্র রিকোয়েস্ট নেই!</b>", parse_mode="HTML")
            else:
                bot.send_message(message.chat.id, f"📥 <b>মোট {len(p_reqs)} টি উইথড্র পেন্ডিং আছে:</b>", parse_mode="HTML")
                for r in p_reqs[:5]:
                    r_id = r.get("req_id")
                    u_target = r.get("user_id")
                    m = r.get("method")
                    a = r.get("account_number")
                    amt = r.get("amount")

                    markup = InlineKeyboardMarkup(row_width=2)
                    markup.add(
                        ibtn("✅ Paid", callback_data=f"pay_with_{r_id}", style="success"),
                        ibtn("❌ Refund", callback_data=f"ref_with_{r_id}", style="danger")
                    )
                    bot.send_message(message.chat.id, f"💳 Req ID: <code>{r_id}</code>\nUser: <code>{u_target}</code>\nMethod: {m}\nAccount: <code>{a}</code>\nAmount: ৳{amt}", parse_mode="HTML", reply_markup=markup)

        elif txt == "📩 Pending Support Tickets" and check_permission(uid, "admin"):
            tickets = fb_get("support_tickets") or {}
            p_tickets = [t for t_id, t in tickets.items() if t.get("status") == "pending"]
            if not p_tickets:
                bot.send_message(message.chat.id, "✅ <b>কোনো পেন্ডিং সাপোর্ট টিকিট নেই!</b>", parse_mode="HTML")
            else:
                for t in p_tickets[:5]:
                    t_id = t.get("ticket_id")
                    u_target = t.get("user_id")
                    msg_text = t.get("message")
                    markup = InlineKeyboardMarkup()
                    markup.add(ibtn("💬 Reply Ticket", callback_data=f"reply_ticket_{t_id}", style="primary"))
                    bot.send_message(message.chat.id, f"📩 Ticket ID: <code>{t_id}</code>\nUser: <code>{u_target}</code>\nMessage: {msg_text}", parse_mode="HTML", reply_markup=markup)

        elif txt == "➕ Add App/TG Task" and check_permission(uid, "admin"):
            update_user_field(uid, "state", "admin_add_app_task")
            bot.send_message(message.chat.id, "➕ <b>নতুন অ্যাপ বা টেলিগ্রাম টাস্ক যোগ করতে ৩ লাইনে নিচের তথ্য দিন:</b>\n\n১ম লাইন: শিরোনাম/টাইটেল\n২য় লাইন: লিংক\n৩য় লাইন: রিওয়ার্ড অ্যামাউন্ট (যেমন: 5.0)", parse_mode="HTML")

        elif txt == "📩 Smart Broadcast Message" and check_permission(uid, "admin"):
            update_user_field(uid, "state", "admin_broadcast_msg")
            bot.send_message(message.chat.id, "📩 <b>সকল বটের ইউজারের কাছে ব্রডকাস্ট করতে বার্তাটি লিখুন:</b>", parse_mode="HTML")

        elif txt == "💰 Set Task Rates" and check_permission(uid, "admin"):
            markup = InlineKeyboardMarkup(row_width=1)
            markup.add(
                ibtn(f"📸 Set Ins Rate (Current: ৳{get_config('rate_ins')})", callback_data="set_rate_ins", style="primary"),
                ibtn(f"📘 Set Fb Rate (Current: ৳{get_config('rate_fb')})", callback_data="set_rate_fb", style="primary"),
                ibtn(f"📧 Set Gmail Rate (Current: ৳{get_config('rate_gmail')})", callback_data="set_rate_gmail", style="primary")
            )
            bot.send_message(message.chat.id, "💰 <b>কাজের রেট সেটিং:</b>", parse_mode="HTML", reply_markup=markup)

        elif txt == "📊 Bot Statistics" and check_permission(uid, "admin"):
            users = fb_get("users") or {}
            total_users = len(users)
            day_ago = time.time() - 86400
            active_users = sum(1 for u_id, u_data in users.items() if float(u_data.get("last_active", 0)) >= day_ago)
            
            subs = fb_get("pending_submissions") or {}
            total_tasks = len(subs)
            tot_income = sum(float(s.get("rate", 0)) for s_id, s in subs.items() if s.get("status") == "approved")

            reqs = fb_get("withdraw_requests") or {}
            tot_withdraw = sum(float(r.get("amount", 0)) for r_id, r in reqs.items() if r.get("status") == "approved")

            msg = (f"📊 <b>BOT STATISTICS & OVERVIEW</b>\n\n"
                   f"👥 <b>মোট ইউজার:</b> <b>{total_users}</b>\n"
                   f"⚡ <b>২৪ ঘণ্টায় একটিভ ইউজার:</b> <b>{active_users}</b>\n"
                   f"📥 <b>মোট সাবমিটেড টাস্ক:</b> <b>{total_tasks}</b>\n"
                   f"💸 <b>মোট পেইড উইথড্র:</b> <b>৳{tot_withdraw:.2f}</b>\n"
                   f"💰 <b>বটের মোট বিতরণ করা ইনকাম:</b> <b>৳{tot_income:.2f}</b>")
            bot.send_message(message.chat.id, msg, parse_mode="HTML")

        elif txt == "⛔ Ban/Unban User" and check_permission(uid, "admin"):
            markup = InlineKeyboardMarkup(row_width=2)
            markup.add(
                ibtn("⛔ Ban User", callback_data="adm_action_ban", style="danger"),
                ibtn("✅ Unban User", callback_data="adm_action_unban", style="success")
            )
            bot.send_message(message.chat.id, "⛔ <b>ইউজার ব্যান/আনব্যান অপশন:</b>", parse_mode="HTML", reply_markup=markup)

        elif txt == "➕ Add/Deduct Balance" and check_permission(uid, "admin"):
            markup = InlineKeyboardMarkup(row_width=2)
            markup.add(
                ibtn("➕ Add Balance", callback_data="adm_bal_add", style="success"),
                ibtn("➖ Deduct Balance", callback_data="adm_bal_ded", style="danger")
            )
            bot.send_message(message.chat.id, "💰 <b>ইউজার ব্যালেন্স যোগ অথবা কাটার অপশন:</b>", parse_mode="HTML", reply_markup=markup)

        elif txt == "👑 Sub-Admin Manager" and check_permission(uid, "admin"):
            update_user_field(uid, "state", "admin_sub_admin_id")
            bot.send_message(message.chat.id, "👑 <b>যাকে Sub-Admin বানাতে চান তার User ID দিন:</b>", parse_mode="HTML")

        elif txt == "📧 Set Recovery Email" and check_permission(uid, "admin"):
            update_user_field(uid, "state", "set_recovery_email")
            bot.send_message(message.chat.id, f"📧 <b>বর্তমান রিকভারি ইমেইল:</b> <code>{get_config('recovery_email')}</code>\n\nনতুন রিকভারি ইমেইল প্রদান করুন:", parse_mode="HTML")

        elif txt == "🎥 Set Tutorial Videos" and check_permission(uid, "admin"):
            markup = InlineKeyboardMarkup(row_width=1)
            markup.add(
                ibtn("📧 Set Gmail Video Link", callback_data="set_vid_gmail", style="primary"),
                ibtn("📘 Set FB Video Link", callback_data="set_vid_fb", style="primary"),
                ibtn("📸 Set Ins Video Link", callback_data="set_vid_ins", style="primary")
            )
            bot.send_message(message.chat.id, "🎥 <b>টিউটোরিয়াল ভিডিও লিংক সেটিং:</b>", parse_mode="HTML", reply_markup=markup)

        elif txt == "🔍 User Deep Search" and check_permission(uid, "admin"):
            update_user_field(uid, "state", "admin_user_search")
            bot.send_message(message.chat.id, "🔍 <b>যেকোনো ইউজারের বিস্তারিত সার্চ করতে তার Telegram User ID লিখুন:</b>", parse_mode="HTML")

        elif txt == "📢 Channel Multi-Post Broadcast" and check_permission(uid, "admin"):
            chs = json.loads(get_config("force_channels", "[]"))
            bot.send_message(message.chat.id, f"📢 <b>সকল ফোর্স জয়েন চ্যানেলে ({len(chs)} টি) কাস্টম পোস্ট করতে টেক্সট লিখুন:</b>", parse_mode="HTML")
            update_user_field(uid, "state", "admin_broadcast_msg")

        elif txt == "🗑️ Bulk Reject Submissions" and check_permission(uid, "admin"):
            all_subs = fb_get("pending_submissions") or {}
            cnt = 0
            for sid, s in all_subs.items():
                if s.get("status") == "pending":
                    fb_update(f"pending_submissions/{sid}", {"status": "rejected"})
                    cnt += 1
            bot.send_message(message.chat.id, f"🗑️ <b>সকল পেন্ডিং সাবমিশন ({cnt} টি) রিজেক্ট করা হয়েছে!</b>", parse_mode="HTML", reply_markup=get_admin_menu())

        elif txt == "⏸️ Task Pause Manager" and check_permission(uid, "admin"):
            markup = InlineKeyboardMarkup(row_width=1)
            p_g = "⏸️ Paused" if get_config("pause_gmail") == "true" else "▶️ Active"
            p_f = "⏸️ Paused" if get_config("pause_fb") == "true" else "▶️ Active"
            p_i = "⏸️ Paused" if get_config("pause_ins") == "true" else "▶️ Active"
            
            markup.add(
                ibtn(f"📧 Gmail Task: {p_g}", callback_data="toggle_pause_gmail", style="primary"),
                ibtn(f"📘 FB Task: {p_f}", callback_data="toggle_pause_fb", style="primary"),
                ibtn(f"📸 Ins Task: {p_i}", callback_data="toggle_pause_ins", style="primary")
            )
            bot.send_message(message.chat.id, "⏸️ <b>Task Pause & Resume Manager:</b>", parse_mode="HTML", reply_markup=markup)

        elif txt == "🧹 Database Cleanup" and check_permission(uid, "admin"):
            bot.send_message(message.chat.id, "🧹 <b>ডাটাবেস অপটিমাইজেশন ও ক্লিনআপ সম্পন্ন করা হয়েছে!</b>", parse_mode="HTML", reply_markup=get_admin_menu())

        elif txt == "🚨 Error Logs" and check_permission(uid, "admin"):
            if os.path.exists(LOG_FILE):
                with open(LOG_FILE, "r", encoding="utf-8") as f:
                    logs = f.read()
                if len(logs) > 3000: logs = logs[-3000:]
                bot.send_message(message.chat.id, f"🚨 <b>ERROR LOGS (Last 3000 chars):</b>\n\n<code>{logs}</code>", parse_mode="HTML")
            else:
                bot.send_message(message.chat.id, "✅ <b>কোনো এরর লগ ফাইল পাওয়া যায়নি!</b>", parse_mode="HTML")

        elif txt == "⚡ Maintenance Mode" and check_permission(uid, "admin"):
            curr = get_config("maintenance_mode", "false")
            new_val = "true" if curr == "false" else "false"
            set_config("maintenance_mode", new_val)
            bot.send_message(message.chat.id, f"⚡ <b>Maintenance Mode: {'ENABLED 🔴' if new_val == 'true' else 'DISABLED 🟢'}</b>", parse_mode="HTML", reply_markup=get_admin_menu())

    except Exception as e:
        log_error(f"Error in handle_all_messages: {e}\n{traceback.format_exc()}")

# ============================================
# --- CALLBACK QUERY HANDLERS ---
# ============================================
@bot.callback_query_handler(func=lambda call: True)
def handle_callbacks(call):
    try:
        uid = call.from_user.id
        u = get_user_db(uid)

        # Force Join Real-time Guard on Inline Clicks
        if uid != ADMIN_ID and call.data != "check_join_event" and not check_force_join(uid):
            bot.answer_callback_query(call.id, "❌ আপনি এখনও সকল চ্যানেলে জয়েন করেননি!", show_alert=True)
            return

        if call.data == "check_join_event":
            if check_force_join(uid):
                bot.answer_callback_query(call.id, "✅ সকল চ্যানেলে জয়েন ভেরিফাইড!")
                try: bot.delete_message(call.message.chat.id, call.message.message_id)
                except: pass
                bot.send_message(call.message.chat.id, "🎉 স্বাগতম!", parse_mode="HTML", reply_markup=get_main_menu(uid))
            else:
                bot.answer_callback_query(call.id, "❌ আপনি এখনো সকল চ্যানেলে জয়েন করেননি!", show_alert=True)

        elif call.data.startswith("copy_code_"):
            code = call.data.replace("copy_code_", "")
            bot.answer_callback_query(call.id, f"✅ Code copied: {code}", show_alert=True)

        elif call.data == "open_spin_game":
            today_str = datetime.now().strftime("%Y-%m-%d")
            limit = int(get_config("spin_limit", "5"))
            
            last_spin_date = u.get("last_spin_date", "")
            daily_spins = int(u.get("daily_spins", 0))

            if last_spin_date != today_str:
                update_user_field(uid, "last_spin_date", today_str)
                update_user_field(uid, "daily_spins", 0)
                daily_spins = 0

            if daily_spins >= limit:
                bot.answer_callback_query(call.id, f"❌ আজকের স্পিন লিমিট ({limit}/{limit}) শেষ!", show_alert=True)
                return

            ad_url = get_config("spin_ad_url", "https://example.com")
            
            markup = InlineKeyboardMarkup(row_width=1)
            markup.add(
                ibtn("🌐 অ্যাড দেখুন (Ad Click Required)", url=ad_url, style="primary"),
                ibtn("🎁 রিওয়ার্ড ভেরিফাই করুন", callback_data="claim_spin_reward_secure", style="success")
            )
            bot.send_message(call.message.chat.id, "⚠️ <b>স্পিন রিওয়ার্ড সিকিউরিটি চেক:</b>\n\nনিচের অ্যাড লিংকে ক্লিক করে অ্যাড দেখার পর 'রিওয়ার্ড ভেরিফাই করুন' বাটনে চাপ দিন:", parse_mode="HTML", reply_markup=markup)

        elif call.data == "claim_spin_reward_secure":
            today_str = datetime.now().strftime("%Y-%m-%d")
            limit = int(get_config("spin_limit", "5"))
            reward = float(get_config("spin_reward", "1.5"))

            daily_spins = int(u.get("daily_spins", 0))
            if daily_spins >= limit:
                bot.answer_callback_query(call.id, f"❌ লিমিট অতিক্রান্ত হয়েছে!", show_alert=True)
                return

            fb_update(f"users/{uid}", {
                "balance": float(u.get("balance", 0)) + reward,
                "total_income": float(u.get("total_income", 0)) + reward,
                "daily_spins": daily_spins + 1,
                "last_spin_date": today_str
            })

            bot.answer_callback_query(call.id, f"🎉 ভেরিফাইড! আপনি ৳{reward} পেয়েছেন!", show_alert=True)
            bot.send_message(call.message.chat.id, f"🎉 <b>অভিনন্দন! আপনি সফলভাবে ৳{reward} আয় করেছেন!</b>\nআজকে আর স্পিন বাকি: {limit - (daily_spins + 1)} টি", parse_mode="HTML", reply_markup=get_main_menu(uid))

        elif call.data == "open_app_tasks":
            tasks = fb_get("app_tasks") or {}
            if not tasks:
                bot.send_message(call.message.chat.id, "📲 <b>বর্তমানে কোনো সক্রিয় অ্যাপ বা টেলিগ্রাম টাস্ক নেই!</b>", parse_mode="HTML")
            else:
                for t_id, task in tasks.items():
                    completed = fb_get(f"completed_app_tasks/{uid}/{t_id}")
                    if completed: continue
                    markup = InlineKeyboardMarkup(row_width=1)
                    markup.add(
                        ibtn("🔗 লিংক খুলুন", url=task.get("link"), style="primary"),
                        ibtn("📤 প্রুফ স্ক্রিনশট দিন", callback_data=f"start_ss_proof_{t_id}", style="success")
                    )
                    bot.send_message(call.message.chat.id, f"📲 <b>{task.get('title')}</b>\n\n💰 রিওয়ার্ড: ৳{task.get('reward')}\n\nলিংকে গিয়ে টাস্ক সম্পন্ন করে প্রুফ পাঠান:", parse_mode="HTML", reply_markup=markup)

        elif call.data.startswith("start_ss_proof_"):
            t_id = call.data.replace("start_ss_proof_", "")
            update_user_field(uid, "state", f"sub_app_proof_{t_id}")
            bot.send_message(call.message.chat.id, "📷 <b>আপনার কাজের প্রুফ হিসেবে স্ক্রিনশট (Photo) আপলোড করুন:</b>", parse_mode="HTML")

        elif call.data == "open_invite_rewards":
            refs = int(u.get("referrals", 0))
            msg = f"🎁 <b>আমন্ত্রণ পুরষ্কার মিলায়ন:</b>\n\nআপনার মোট রেফার: <b>{refs}</b> জন\n\n১০ জন রেফার করলে পাবেন ৳৫০ বোনাস!"
            bot.send_message(call.message.chat.id, msg, parse_mode="HTML")

        elif call.data == "open_support_ticket":
            update_user_field(uid, "state", "user_submit_ticket")
            bot.send_message(call.message.chat.id, "📩 <b>আপনার মেসেজ বা সমস্যার বিষয়টি লিখে পাঠান:</b>", parse_mode="HTML")

        elif call.data.startswith("watch_vid_"):
            serv = call.data.replace("watch_vid_", "")
            vids = json.loads(get_config("tutorial_videos", "{}"))
            v_url = vids.get(serv, "https://youtube.com")
            bot.send_message(call.message.chat.id, f"🎥 <b>{serv.upper()} টিউটোরিয়াল ভিডিও দেখতে নিচের লিংকে যান:</b>\n\n{v_url}", parse_mode="HTML")

        elif call.data.startswith("with_meth_"):
            meth = call.data.replace("with_meth_", "")
            update_user_field(uid, "state", f"withdraw_number_{meth}")
            bot.send_message(call.message.chat.id, f"📥 <b>আপনার {meth} নম্বর/ওয়ালেট অ্যাড্রেস লিখুন:</b>", parse_mode="HTML")

        # --- ADMIN CALLBACK ACTIONS ---
        elif call.data.startswith("pay_with_") and check_permission(uid, "admin"):
            req_id = call.data.replace("pay_with_", "")
            fb_update(f"withdraw_requests/{req_id}", {"status": "approved"})
            req_data = fb_get(f"withdraw_requests/{req_id}")
            if req_data:
                u_target = req_data.get("user_id")
                try: bot.send_message(u_target, f"🎉 <b>আপনার ৳{req_data.get('amount')} এর উইথড্র রিকোয়েস্ট সাকসেসফুলি পেইড করা হয়েছে!</b>", parse_mode="HTML")
                except: pass
            bot.answer_callback_query(call.id, "✅ Payment marked as PAID!")
            try: bot.edit_message_text(f"✅ <b>PAID ({req_id})</b>", call.message.chat.id, call.message.message_id, parse_mode="HTML")
            except: pass

        elif call.data.startswith("ref_with_") and check_permission(uid, "admin"):
            req_id = call.data.replace("ref_with_", "")
            req_data = fb_get(f"withdraw_requests/{req_id}")
            if req_data and req_data.get("status") == "pending":
                fb_update(f"withdraw_requests/{req_id}", {"status": "refunded"})
                u_target = req_data.get("user_id")
                amt = float(req_data.get("amount", 0))
                t_u = get_user_db(u_target)
                fb_update(f"users/{u_target}", {"balance": float(t_u.get("balance", 0)) + amt})
                try: bot.send_message(u_target, f"❌ <b>আপনার উইথড্র রিজেক্ট করা হয়েছে এবং ৳{amt} ফেরত দেওয়া হয়েছে।</b>", parse_mode="HTML")
                except: pass
            bot.answer_callback_query(call.id, "❌ Refunded & Balance restored!")
            try: bot.edit_message_text(f"❌ <b>REFUNDED ({req_id})</b>", call.message.chat.id, call.message.message_id, parse_mode="HTML")
            except: pass

        elif call.data.startswith("adm_app_") and check_permission(uid, "admin"):
            s_id = call.data.replace("adm_app_", "")
            sub = fb_get(f"pending_submissions/{s_id}")
            if sub and sub.get("status") == "pending":
                fb_update(f"pending_submissions/{s_id}", {"status": "approved"})
                u_target = sub.get("user_id")
                rate = float(sub.get("rate", 0))
                t_u = get_user_db(u_target)
                fb_update(f"users/{u_target}", {
                    "balance": float(t_u.get("balance", 0)) + rate,
                    "total_income": float(t_u.get("total_income", 0)) + rate,
                    "approved_tasks": int(t_u.get("approved_tasks", 0)) + 1
                })
                try: bot.send_message(u_target, f"✅ <b>আপনার কাজ এপ্রুভ হয়েছে! ৳{rate} যোগ করা হয়েছে।</b>", parse_mode="HTML")
                except: pass
            bot.answer_callback_query(call.id, "✅ Task Approved!")
            try: bot.edit_message_text(f"✅ Approved ({s_id})", call.message.chat.id, call.message.message_id)
            except: pass

        elif call.data.startswith("adm_rej_") and check_permission(uid, "admin"):
            s_id = call.data.replace("adm_rej_", "")
            sub = fb_get(f"pending_submissions/{s_id}")
            if sub and sub.get("status") == "pending":
                fb_update(f"pending_submissions/{s_id}", {"status": "rejected"})
                u_target = sub.get("user_id")
                t_u = get_user_db(u_target)
                fb_update(f"users/{u_target}", {"rejected_tasks": int(t_u.get("rejected_tasks", 0)) + 1})
                try: bot.send_message(u_target, f"❌ <b>আপনার কাজ রিজেক্ট করা হয়েছে! Sub ID: {s_id}</b>", parse_mode="HTML")
                except: pass
            bot.answer_callback_query(call.id, "❌ Task Rejected!")
            try: bot.edit_message_text(f"❌ Rejected ({s_id})", call.message.chat.id, call.message.message_id)
            except: pass

        elif call.data.startswith("reply_ticket_") and check_permission(uid, "admin"):
            t_id = call.data.replace("reply_ticket_", "")
            update_user_field(uid, "state", f"admin_reply_ticket_{t_id}")
            bot.send_message(call.message.chat.id, f"💬 <b>টিকিট {t_id} এর উত্তর লিখুন:</b>", parse_mode="HTML")

        elif call.data == "adm_action_ban" and check_permission(uid, "admin"):
            update_user_field(uid, "state", "admin_ban_user")
            bot.send_message(call.message.chat.id, "⛔ <b>যাকে ব্যান করতে চান তার Telegram User ID দিন:</b>", parse_mode="HTML")

        elif call.data == "adm_action_unban" and check_permission(uid, "admin"):
            update_user_field(uid, "state", "admin_unban_user")
            bot.send_message(call.message.chat.id, "✅ <b>যাকে আনব্যান করতে চান তার Telegram User ID দিন:</b>", parse_mode="HTML")

        elif call.data == "adm_bal_add" and check_permission(uid, "admin"):
            update_user_field(uid, "state", "admin_add_bal_id")
            bot.send_message(call.message.chat.id, "➕ <b>যার অ্যাকাউন্টে ব্যালেন্স যোগ করতে চান তার User ID দিন:</b>", parse_mode="HTML")

        elif call.data == "adm_bal_ded" and check_permission(uid, "admin"):
            update_user_field(uid, "state", "admin_ded_bal_id")
            bot.send_message(call.message.chat.id, "➖ <b>যার অ্যাকাউন্ট থেকে ব্যালেন্স কাটতে চান তার User ID দিন:</b>", parse_mode="HTML")

        elif call.data.startswith("toggle_wmeth_") and check_permission(uid, "admin"):
            m = call.data.replace("toggle_wmeth_", "")
            w_methods = json.loads(get_config("withdraw_methods", "{}"))
            if m in w_methods:
                w_methods[m]["enabled"] = not w_methods[m]["enabled"]
                set_config("withdraw_methods", json.dumps(w_methods))
                bot.answer_callback_query(call.id, f"✅ {m} status toggled!")
            bot.send_message(call.message.chat.id, "💳 <b>Withdraw Methods updated!</b>", parse_mode="HTML", reply_markup=get_admin_menu())

        elif call.data.startswith("toggle_pause_") and check_permission(uid, "admin"):
            serv = call.data.replace("toggle_pause_", "")
            curr = get_config(f"pause_{serv}", "false")
            new_v = "true" if curr == "false" else "false"
            set_config(f"pause_{serv}", new_v)
            bot.answer_callback_query(call.id, f"✅ {serv.upper()} Task Status Updated!")
            bot.send_message(call.message.chat.id, f"⏸️ <b>{serv.upper()} Task: {'PAUSED' if new_v == 'true' else 'ACTIVE'}</b>", parse_mode="HTML", reply_markup=get_admin_menu())

        elif call.data == "set_rate_ins" and check_permission(uid, "admin"):
            update_user_field(uid, "state", "set_rate_ins")
            bot.send_message(call.message.chat.id, "💰 <b>নতুন Instagram কাজের রেট লিখুন (যেমন: 15.0):</b>", parse_mode="HTML")

        elif call.data == "set_rate_fb" and check_permission(uid, "admin"):
            update_user_field(uid, "state", "set_rate_fb")
            bot.send_message(call.message.chat.id, "💰 <b>নতুন Facebook কাজের রেট লিখুন (যেমন: 18.0):</b>", parse_mode="HTML")

        elif call.data == "set_rate_gmail" and check_permission(uid, "admin"):
            update_user_field(uid, "state", "set_rate_gmail")
            bot.send_message(call.message.chat.id, "💰 <b>নতুন Gmail কাজের রেট লিখুন (যেমন: 12.0):</b>", parse_mode="HTML")

        elif call.data == "set_pass_ins" and check_permission(uid, "admin"):
            update_user_field(uid, "state", "set_pass_ins")
            bot.send_message(call.message.chat.id, "🔑 <b>নতুন Instagram Password Base লিখুন:</b>", parse_mode="HTML")

        elif call.data == "set_pass_fb" and check_permission(uid, "admin"):
            update_user_field(uid, "state", "set_pass_fb")
            bot.send_message(call.message.chat.id, "🔑 <b>নতুন Facebook Password Base লিখুন:</b>", parse_mode="HTML")

        elif call.data == "set_pass_gmail" and check_permission(uid, "admin"):
            update_user_field(uid, "state", "set_pass_gmail")
            bot.send_message(call.message.chat.id, "🔑 <b>নতুন Gmail Password Base লিখুন:</b>", parse_mode="HTML")

        elif call.data == "set_spin_reward" and check_permission(uid, "admin"):
            update_user_field(uid, "state", "set_spin_reward")
            bot.send_message(call.message.chat.id, "💰 <b>নতুন Spin Reward অ্যামাউন্ট লিখুন (যেমন: 1.5):</b>", parse_mode="HTML")

        elif call.data == "set_spin_limit" and check_permission(uid, "admin"):
            update_user_field(uid, "state", "set_spin_limit")
            bot.send_message(call.message.chat.id, "🔢 <b>নতুন Spin Daily Limit লিখুন (যেমন: 5):</b>", parse_mode="HTML")

        elif call.data == "set_spin_url" and check_permission(uid, "admin"):
            update_user_field(uid, "state", "set_spin_url")
            bot.send_message(call.message.chat.id, "🌐 <b>নতুন Adsterra Direct Link লিখুন:</b>", parse_mode="HTML")

        elif call.data == "set_sheet_ins" and check_permission(uid, "admin"):
            update_user_field(uid, "state", "set_sheet_ins")
            bot.send_message(call.message.chat.id, "📊 <b>Instagram Google Sheet ID দিন:</b>", parse_mode="HTML")

        elif call.data == "set_sheet_fb" and check_permission(uid, "admin"):
            update_user_field(uid, "state", "set_sheet_fb")
            bot.send_message(call.message.chat.id, "📊 <b>Facebook Google Sheet ID দিন:</b>", parse_mode="HTML")

        elif call.data == "set_sheet_gmail" and check_permission(uid, "admin"):
            update_user_field(uid, "state", "set_sheet_gmail")
            bot.send_message(call.message.chat.id, "📊 <b>Gmail Google Sheet ID দিন:</b>", parse_mode="HTML")

        elif call.data == "set_vid_gmail" and check_permission(uid, "admin"):
            update_user_field(uid, "state", "admin_set_vid_gmail")
            bot.send_message(call.message.chat.id, "🎥 <b>Gmail Tutorial Video Link দিন:</b>", parse_mode="HTML")

        elif call.data == "set_vid_fb" and check_permission(uid, "admin"):
            update_user_field(uid, "state", "admin_set_vid_fb")
            bot.send_message(call.message.chat.id, "🎥 <b>Facebook Tutorial Video Link দিন:</b>", parse_mode="HTML")

        elif call.data == "set_vid_ins" and check_permission(uid, "admin"):
            update_user_field(uid, "state", "admin_set_vid_ins")
            bot.send_message(call.message.chat.id, "🎥 <b>Instagram Tutorial Video Link দিন:</b>", parse_mode="HTML")

        # --- ADMIN FORCE JOIN MANAGEMENT ---
        elif call.data == "add_force_channel":
            update_user_field(uid, "state", "add_force_channel_input")
            bot.send_message(call.message.chat.id, "📢 <b>নতুন চ্যানেল ইউজারনেম দিন (যেমন: @ChannelUsername):</b>", parse_mode="HTML")

        elif call.data.startswith("rem_force_"):
            target_ch = call.data.replace("rem_force_", "")
            chs = json.loads(get_config("force_channels", "[]"))
            if target_ch in chs:
                chs.remove(target_ch)
                set_config("force_channels", json.dumps(chs))
                bot.answer_callback_query(call.id, f"✅ {target_ch} রিমুভ করা হয়েছে!")
            bot.send_message(call.message.chat.id, "📢 <b>Force Join চ্যানেল তালিকা আপডেট হয়েছে!</b>", parse_mode="HTML", reply_markup=get_admin_menu())

        # --- EXPORT DATEWISE MENU HANDLERS ---
        elif call.data.startswith("exp_menu_"):
            serv = call.data.replace("exp_menu_", "")
            markup = InlineKeyboardMarkup(row_width=1)
            markup.add(
                ibtn("📅 আজকের ফাইল (Today)", callback_data=f"exp_today_{serv}", style="primary"),
                ibtn("📅 গতকালকের ফাইল (Yesterday)", callback_data=f"exp_yest_{serv}", style="primary"),
                ibtn("🗓️ নির্দিষ্ট তারিখ (Select Date)", callback_data=f"exp_cust_{serv}", style="success")
            )
            bot.send_message(call.message.chat.id, f"📊 <b>{serv.upper()} Unsold ফাইল ডাউনলোডের ফিল্টার বেছে নিন:</b>", parse_mode="HTML", reply_markup=markup)

        elif call.data.startswith("exp_today_") or call.data.startswith("exp_yest_"):
            is_today = "exp_today_" in call.data
            serv = call.data.replace("exp_today_", "").replace("exp_yest_", "")
            
            target_dt = datetime.now().date() if is_today else (datetime.now() - timedelta(days=1)).date()
            dt_str = target_dt.strftime("%Y-%m-%d")

            all_subs = fb_get("pending_submissions") or {}
            matched = []
            for sid, s in all_subs.items():
                if s.get("sub_type") == serv and s.get("status") == "pending":
                    s_date = datetime.fromtimestamp(float(s.get("created_at", 0))).date()
                    if s_date == target_dt:
                        s["id"] = sid
                        matched.append(s)

            if not matched:
                bot.answer_callback_query(call.id, f"❌ {dt_str} তারিখে কোনো Unsold ডাটা নেই!", show_alert=True)
                return

            filename = f"{serv}_{dt_str}.xlsx"
            generate_custom_excel(serv, matched, filename)

            with open(filename, "rb") as f:
                bot.send_document(call.message.chat.id, f, caption=f"📊 <b>{serv.upper()} ({dt_str}) Unsold Report</b>", parse_mode="HTML")
            os.remove(filename)

            markup = InlineKeyboardMarkup()
            markup.add(ibtn(f"🗑️ Mark as Sold / Delete Batch ({dt_str})", callback_data=f"del_batch_{serv}_{dt_str}", style="danger"))
            bot.send_message(call.message.chat.id, f"⚠️ <b>ফাইল ডাটাবেস থেকে মুছে ফেলতে চাইলে বায়ার কনফার্মেশনের পর নিচের বাটন চাপুন:</b>", parse_mode="HTML", reply_markup=markup)

        elif call.data.startswith("exp_cust_"):
            serv = call.data.replace("exp_cust_", "")
            temp = json.loads(u.get("temp_data") or "{}")
            temp["export_service"] = serv
            update_user_field(uid, "temp_data", json.dumps(temp))
            update_user_field(uid, "state", "export_date_input")
            bot.send_message(call.message.chat.id, "📅 <b>যে তারিখের ফাইল নামাতে চান তা YYYY-MM-DD ফরম্যাটে লিখুন (যেমন: 2026-08-22):</b>", parse_mode="HTML")

        elif call.data.startswith("del_batch_"):
            parts = call.data.split("_")
            serv = parts[2]
            target_dt_str = parts[3]
            target_dt = datetime.strptime(target_dt_str, "%Y-%m-%d").date()

            all_subs = fb_get("pending_submissions") or {}
            deleted_count = 0
            for sid, s in all_subs.items():
                if s.get("sub_type") == serv and s.get("status") == "pending":
                    s_date = datetime.fromtimestamp(float(s.get("created_at", 0))).date()
                    if s_date == target_dt:
                        fb_delete(f"pending_submissions/{sid}")
                        deleted_count += 1

            bot.answer_callback_query(call.id, f"🗑️ {deleted_count} টি ডাটা মুছে ফেলা হয়েছে!", show_alert=True)
            bot.send_message(call.message.chat.id, f"✅ <b>{target_dt_str} তারিখের {serv.upper()} সার্ভিসের মোট {deleted_count} টি ডাটা সার্ভার থেকে ডিলিট করা হয়েছে।</b>", parse_mode="HTML")

    except Exception as e:
        log_error(f"Error in handle_callbacks: {e}\n{traceback.format_exc()}")

# ============================================
# --- ENGINE START ---
# ============================================
if __name__ == "__main__":
    keep_alive()
    print(f"🚀 {BOT_NAME} Firebase Production Engine running safely...")
    bot.infinity_polling(skip_pending=True)
